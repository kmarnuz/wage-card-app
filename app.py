"""
Wage Card Management System — All-in-One App
Run: python3 app.py
Open: http://localhost:8000
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend', 'src'))

from frontend_html import FRONTEND_HTML

from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import uuid
import io
from datetime import datetime

from services.calculation_engine import WageCalculationEngine, WageInput
from services.database import db
from services.excel_export import export_wage_cards_to_excel
from config.statutory_data import get_ptax_slabs, get_lwf_config, PTAX_SLABS, LWF_RATES

app = FastAPI(title="Wage Card Management System", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
engine = WageCalculationEngine()

# --- Parity Groups Configuration ---
import json as _json
PARITY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'parity_groups.json')
print(f"📋 Parity config path: {PARITY_FILE}, exists: {os.path.exists(PARITY_FILE)}")

def load_parity_groups():
    """Load parity group config from JSON file."""
    if os.path.exists(PARITY_FILE):
        with open(PARITY_FILE, 'r') as f:
            data = _json.load(f)
            groups = data.get('parity_groups', [])
            return groups
    print(f"⚠️ Parity file not found: {PARITY_FILE}")
    return []

def get_parity_group_for_site(site_code, entity=None):
    """Find which parity group a site belongs to (if any)."""
    groups = load_parity_groups()
    for group in groups:
        if site_code.upper() in [s.upper() for s in group.get('sites', [])]:
            if entity and group.get('entity') and group['entity'].upper() != entity.upper():
                continue
            return group
    return None

def enforce_parity(all_cards, engine_ref, get_ptax_fn, get_lwf_fn):
    """
    Enforce parity across parity groups.
    For each group, find the HIGHEST Gross per (short_bt, tenure_years)
    and apply that Gross to all sites in the group.
    Returns number of cards updated.
    """
    groups = load_parity_groups()
    if not groups:
        return 0

    updated_count = 0

    for group in groups:
        group_sites = [s.upper() for s in group.get('sites', [])]
        group_entity = group.get('entity', '').upper() if group.get('entity') else None

        # Find all cards belonging to this parity group (non-PT only)
        group_cards = []
        for card in all_cards:
            site = card.get('site_codes', '').upper()
            entity = card.get('entity', '').upper()
            if site in group_sites and not card.get('is_pt'):
                if group_entity and entity != group_entity:
                    continue
                group_cards.append(card)

        if not group_cards:
            continue

        # Group by (short_bt, tenure_years) to find max Gross per role+tenure
        from collections import defaultdict
        role_tenure_groups = defaultdict(list)
        for card in group_cards:
            key = (card.get('short_bt', ''), card.get('tenure_years', 0))
            role_tenure_groups[key].append(card)

        # For each role+tenure combo, enforce highest Gross with IDENTICAL split
        for (bt, tenure), cards_in_combo in role_tenure_groups.items():
            # Find the reference card: highest Gross, and on a Gross tie prefer the
            # higher-MW site (its split is MW-driven / structurally correct — e.g. more
            # in LTA). Lower-MW sites in the group then copy this exact split verbatim,
            # even though their own MW is lower (they become over-compliant). This keeps
            # every site in the parity group on an identical component split.
            ref_card = max(
                cards_in_combo,
                key=lambda c: (c.get('gross', 0), c.get('minimum_wage', 0))
            )
            max_gross = ref_card.get('gross', 0)
            ref_basic = ref_card.get('basic', 0)
            ref_flexi = ref_card.get('flexi', 0)
            ref_lta = ref_card.get('lta', 0)
            ref_hra = ref_card.get('hra', 0)
            ref_conveyance = ref_card.get('conveyance', 0)

            for card in cards_in_combo:
                if card is ref_card:
                    continue

                state = card.get('state', '')

                # Copy the EXACT reference split verbatim (Gross + component split
                # identical across the whole parity group, regardless of this site's
                # own MW). Lower-MW sites become over-compliant, which is intended.
                use_basic, use_flexi = ref_basic, ref_flexi
                use_lta, use_hra, use_conveyance = ref_lta, ref_hra, ref_conveyance

                # Update only if the card's current split differs from the reference split
                needs_update = (
                    card.get('gross', 0) < max_gross
                    or card.get('basic', 0) != use_basic
                    or card.get('flexi', 0) != use_flexi
                    or card.get('lta', 0) != use_lta
                    or card.get('hra', 0) != use_hra
                    or card.get('conveyance', 0) != use_conveyance
                )
                if not needs_update:
                    continue

                wage_input = WageInput(
                    state=state, city=card.get('city', ''),
                    site_code=card.get('site_codes', ''),
                    entity=card.get('entity', ''), mw_zone=card.get('mw_zone', ''),
                    region=card.get('region', ''), mw_category=card.get('mw_category', ''),
                    business_title=card.get('short_bt', ''), short_bt=card.get('short_bt', ''),
                    weekly_hours=card.get('weekly_hours', 45),
                    daily_hours=card.get('daily_hours', 9),
                    minimum_wage=card.get('minimum_wage', 0),
                    mw_effective_date=card.get('mw_effective_date', ''),
                    basic=use_basic, flexi=use_flexi, lta=use_lta,
                    hra=use_hra, conveyance=use_conveyance,
                    attendance_incentive=card.get('attendance_incentive', 0),
                    tenure_years=card.get('tenure_years', 0),
                )

                ptax = get_ptax_fn(state)
                lwf = get_lwf_fn(state)
                result = engine_ref.calculate(wage_input, ptax, lwf)

                # Update card with identical components from reference
                card.update({
                    'basic': result.basic, 'flexi': result.flexi, 'lta': result.lta,
                    'hra': result.hra, 'conveyance': result.conveyance, 'gross': result.gross,
                    'per_hour_ot_total': result.per_hour_ot_total,
                    'per_hour_ot_included': result.per_hour_ot_included,
                    'per_hour_ot_balance': result.per_hour_ot_balance,
                    'pf_employee': result.pf_employee, 'esic_employee': result.esic_employee,
                    'gross_deductions': result.gross_deductions, 'net_salary': result.net_salary,
                    'pf_employer': result.pf_employer, 'esic_employer': result.esic_employer,
                    'ctc': result.ctc,
                    'ot_default': result.per_hour_ot_total * 22 if card.get('weekly_hours', 45) == 40 else 0,
                    'total_remuneration': result.total_remuneration,
                    'included_wages': result.included_wages,
                    'included_pct': round(result.included_pct, 4),
                    'excluded_wages': result.excluded_wages,
                    'cap_50_amount': result.cap_50_amount,
                    'cap_50_met': result.cap_50_met, 'mw_compliant': result.mw_compliant,
                })

                db.put_wage_card(card, skip_save=True)
                updated_count += 1

    return updated_count

# Audit trail storage
import shutil
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
AUDIT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'audit_trail.json')
os.makedirs(UPLOADS_DIR, exist_ok=True)

# Upload password protection
PASSWORD_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'password.txt')

def get_upload_password():
    try:
        if os.path.exists(PASSWORD_FILE):
            with open(PASSWORD_FILE, 'r') as f:
                return f.read().strip()
    except:
        pass
    return "CTK@2026"  # default

def set_upload_password(new_password):
    with open(PASSWORD_FILE, 'w') as f:
        f.write(new_password)

def load_audit_trail():
    try:
        if os.path.exists(AUDIT_FILE):
            with open(AUDIT_FILE, 'r') as f:
                return json.load(f)
    except:
        pass
    return []

def save_audit_entry(action, user="System", filename="", details=""):
    trail = load_audit_trail()
    entry = {
        "id": str(uuid.uuid4()),
        "timestamp": datetime.utcnow().isoformat(),
        "action": action,
        "user": user,
        "filename": filename,
        "details": details,
    }
    trail.append(entry)
    with open(AUDIT_FILE, 'w') as f:
        json.dump(trail, f)
    return entry

import json


# ============================================================
# REQUEST MODELS
# ============================================================

class WageCardRequest(BaseModel):
    entity: str = "AMZL"
    state: str
    state_code: str
    city: str
    mw_zone: str = "A"
    region: str = "NA"
    mw_category: str
    business_title: str
    short_bt: str
    site_codes: str
    tenure_years: int = 0
    weekly_hours: float = 45.0
    daily_hours: float = 9.0
    monthly_ot_limit: Optional[float] = None
    minimum_wage: float
    mw_effective_date: str
    basic: float
    flexi: float = 0.0
    lta: float = 0.0
    hra: float = 0.0
    conveyance: float = 0.0
    attendance_incentive: float = 0.0
    nsa_amount: float = 0.0
    ot_hours: float = 0.0

class AutoSplitRequest(BaseModel):
    target_gross: float
    minimum_wage: float
    state: str
    hra_applicable: bool = False

class MWUpdateRequest(BaseModel):
    state: str
    city: str
    mw_zone: str = "A"
    skill_category: str
    new_mw_amount: float
    effective_date: str
    notification_ref: str = ""
    auto_increase_gross: bool = False

class CalculateRequest(BaseModel):
    state: str
    weekly_hours: float = 45.0
    minimum_wage: float = 0.0
    basic: float
    flexi: float = 0.0
    lta: float = 0.0
    hra: float = 0.0
    conveyance: float = 0.0
    attendance_incentive: float = 0.0
    nsa_amount: float = 0.0


# ============================================================
# API ENDPOINTS
# ============================================================

@app.post("/api/wage-cards")
def create_wage_card(request: WageCardRequest):
    card_id = str(uuid.uuid4())
    wage_input = WageInput(
        state=request.state, city=request.city, site_code=request.site_codes,
        entity=request.entity, mw_zone=request.mw_zone, region=request.region,
        mw_category=request.mw_category, business_title=request.business_title,
        short_bt=request.short_bt, weekly_hours=request.weekly_hours,
        daily_hours=request.daily_hours, minimum_wage=request.minimum_wage,
        mw_effective_date=request.mw_effective_date, basic=request.basic,
        flexi=request.flexi, lta=request.lta, hra=request.hra,
        conveyance=request.conveyance, attendance_incentive=request.attendance_incentive,
        nsa_amount=request.nsa_amount, ot_hours=request.ot_hours,
        tenure_years=request.tenure_years,
    )
    ptax = get_ptax_slabs(request.state)
    lwf = get_lwf_config(request.state)
    result = engine.calculate(wage_input, ptax, lwf)
    now = datetime.utcnow().isoformat()

    card = {
        "id": card_id, "entity": request.entity, "state": request.state,
        "state_code": request.state_code, "city": request.city,
        "mw_zone": request.mw_zone, "region": request.region,
        "mw_category": request.mw_category, "business_title": request.business_title,
        "short_bt": request.short_bt, "site_codes": request.site_codes,
        "tenure_years": request.tenure_years, "weekly_hours": request.weekly_hours,
        "daily_hours": request.daily_hours, "monthly_ot_limit": request.monthly_ot_limit,
        "minimum_wage": request.minimum_wage, "mw_effective_date": request.mw_effective_date,
        "basic": result.basic, "flexi": result.flexi, "lta": result.lta,
        "hra": result.hra, "conveyance": result.conveyance, "gross": result.gross,
        "per_hour_ot_total": result.per_hour_ot_total,
        "per_hour_ot_included": result.per_hour_ot_included,
        "per_hour_ot_balance": result.per_hour_ot_balance,
        "pf_employee": result.pf_employee, "esic_employee": result.esic_employee,
        "pt_employee": "As applicable", "lwf_employee": "As applicable",
        "gross_deductions": result.gross_deductions, "net_salary": result.net_salary,
        "pf_employer": result.pf_employer, "esic_employer": result.esic_employer,
        "lwf_employer": "As applicable", "ctc": result.ctc,
        "ot_default": result.ot_default, "nsa": result.nsa,
        "attendance_incentive": result.attendance_incentive,
        "total_remuneration": result.total_remuneration,
        "included_wages": result.included_wages,
        "included_pct": round(result.included_pct, 4),
        "excluded_wages": result.excluded_wages,
        "cap_50_amount": result.cap_50_amount,
        "cap_50_met": result.cap_50_met, "mw_compliant": result.mw_compliant,
        "created_at": now, "updated_at": now,
    }
    db.put_wage_card(card)
    db.put_audit_entry({"entity_type": "wage_card", "entity_id": card_id, "action": "create", "timestamp": now})
    return card

# Fields that should always be whole numbers (no decimals)
MONETARY_FIELDS = {'basic', 'flexi', 'lta', 'hra', 'conveyance', 'gross', 'minimum_wage',
                   'per_hour_ot_total', 'per_hour_ot_included', 'per_hour_ot_balance',
                   'pf_employee', 'esic_employee', 'gross_deductions', 'net_salary',
                   'pf_employer', 'esic_employer', 'ctc', 'ot_default', 'nsa',
                   'attendance_incentive', 'total_remuneration', 'included_wages',
                   'excluded_wages', 'cap_50_amount', 'hol_wage', 'old_ot', 'old_hol',
                   'bal_pay_ot', 'bal_pay_hol'}

# Entities+States where Holiday Wage columns should be blank
BLANK_HOL_ENTITIES_STATES = {('INFC', 'MH'), ('INFC', 'GJ'), ('GSF HUB', 'MH'), ('GSF HUB', 'GJ')}

def round_card(card):
    """Round all monetary fields to whole numbers and blank hol wage for INFC/GSF HUB MH/GJ."""
    for key in MONETARY_FIELDS:
        if key in card and isinstance(card[key], (int, float)):
            card[key] = round(card[key])
    # Blank Holiday Wage for INFC/GSF HUB in MH/GJ
    entity = card.get('entity', '').upper()
    state = card.get('state', '').upper()
    if (entity, state) in BLANK_HOL_ENTITIES_STATES:
        card['hol_wage'] = ''
        card['old_hol'] = ''
        card['bal_pay_hol'] = ''
    return card

@app.get("/api/wage-cards")
def list_wage_cards(state: Optional[str] = None, city: Optional[str] = None,
                    business_title: Optional[str] = None, tenure_years: Optional[int] = None):
    filters = {}
    if state: filters["state"] = state
    if city: filters["city"] = city
    if business_title: filters["business_title"] = business_title
    if tenure_years is not None: filters["tenure_years"] = tenure_years
    cards = db.list_wage_cards(filters)
    cards = [round_card(c) for c in cards]
    return {"count": len(cards), "items": cards}

@app.get("/api/wage-cards/{card_id}")
def get_wage_card(card_id: str):
    card = db.get_wage_card(card_id)
    if not card: raise HTTPException(404, "Wage card not found")
    return card

@app.delete("/api/wage-cards/clear-all")
def clear_all_wage_cards(password: str = Form("")):
    """Delete all wage cards. Requires password."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password.")
    db._memory_store["wage_cards"] = {}
    db.save()
    return {"status": "cleared"}

@app.delete("/api/wage-cards/{card_id}")
def delete_wage_card(card_id: str):
    if not db.delete_wage_card(card_id): raise HTTPException(404, "Not found")
    db.put_audit_entry({"entity_type": "wage_card", "entity_id": card_id, "action": "delete", "timestamp": datetime.utcnow().isoformat()})
    return {"status": "deleted"}

@app.post("/api/calculate")
def calculate_wage(request: CalculateRequest):
    wage_input = WageInput(state=request.state, city="", site_code="",
        weekly_hours=request.weekly_hours, minimum_wage=request.minimum_wage,
        basic=request.basic, flexi=request.flexi, lta=request.lta,
        hra=request.hra, conveyance=request.conveyance,
        attendance_incentive=request.attendance_incentive, nsa_amount=request.nsa_amount)
    result = engine.calculate(wage_input, get_ptax_slabs(request.state), get_lwf_config(request.state))
    return {"gross": result.gross, "per_hour_ot_total": result.per_hour_ot_total,
            "per_hour_ot_included": result.per_hour_ot_included, "per_hour_ot_balance": result.per_hour_ot_balance,
            "pf_employee": result.pf_employee, "esic_employee": result.esic_employee,
            "gross_deductions": result.gross_deductions, "net_salary": result.net_salary,
            "pf_employer": result.pf_employer, "esic_employer": result.esic_employer,
            "ctc": result.ctc, "included_wages": result.included_wages,
            "mw_compliant": result.mw_compliant, "cap_50_met": result.cap_50_met,
            "total_remuneration": result.total_remuneration}

@app.post("/api/auto-split")
def auto_split(request: AutoSplitRequest):
    return engine.auto_split_for_mw(request.target_gross, request.minimum_wage, request.state, request.hra_applicable)

@app.post("/api/minimum-wages/apply")
def apply_mw_update(request: MWUpdateRequest):
    all_cards = db.list_wage_cards()
    updated, failed = [], []
    now = datetime.utcnow().isoformat()
    for card in all_cards:
        if (card["state"].upper() != request.state.upper() or
            card["city"].upper() != request.city.upper() or
            card["mw_category"].upper() != request.skill_category.upper()): continue
        if card.get("included_wages", 0) >= request.new_mw_amount: continue
        split = engine.auto_split_for_mw(card["gross"], request.new_mw_amount, request.state, card.get("hra", 0) > 0)
        if not split["mw_compliant"] and request.auto_increase_gross:
            new_target = request.new_mw_amount + card.get("hra", 0) + card.get("conveyance", 0)
            split = engine.auto_split_for_mw(new_target, request.new_mw_amount, request.state, card.get("hra", 0) > 0)
        if not split["mw_compliant"]:
            failed.append({"card_id": card["id"], "short_bt": card["short_bt"], "reason": "Cannot meet MW"})
            continue
        wage_input = WageInput(state=card["state"], city=card["city"], site_code="",
            weekly_hours=card.get("weekly_hours", 45), minimum_wage=request.new_mw_amount,
            basic=split["basic"], flexi=split["flexi"], lta=split["lta"],
            hra=split["hra"], conveyance=split["conveyance"],
            attendance_incentive=card.get("attendance_incentive", 0), nsa_amount=card.get("nsa", 0))
        r = engine.calculate(wage_input, get_ptax_slabs(card["state"]), get_lwf_config(card["state"]))
        card.update({"basic": split["basic"], "flexi": split["flexi"], "lta": split["lta"],
            "hra": split["hra"], "conveyance": split["conveyance"], "gross": r.gross,
            "minimum_wage": request.new_mw_amount, "mw_effective_date": request.effective_date,
            "per_hour_ot_total": r.per_hour_ot_total, "per_hour_ot_included": r.per_hour_ot_included,
            "per_hour_ot_balance": r.per_hour_ot_balance, "pf_employee": r.pf_employee,
            "esic_employee": r.esic_employee, "gross_deductions": r.gross_deductions,
            "net_salary": r.net_salary, "pf_employer": r.pf_employer, "esic_employer": r.esic_employer,
            "ctc": r.ctc, "total_remuneration": r.total_remuneration, "included_wages": r.included_wages,
            "included_pct": round(r.included_pct, 4), "excluded_wages": r.excluded_wages,
            "cap_50_amount": r.cap_50_amount, "cap_50_met": r.cap_50_met,
            "mw_compliant": r.mw_compliant, "updated_at": now})
        db.put_wage_card(card)
        updated.append(card["id"])
    return {"total_updated": len(updated), "total_failed": len(failed), "failed_cards": failed}

@app.get("/api/wage-cards/export/excel")
def export_excel(state: Optional[str] = None):
    filters = {"state": state} if state else {}
    cards = db.list_wage_cards(filters)
    if not cards: raise HTTPException(404, "No cards to export")
    cards = [round_card(c) for c in cards]
    excel_bytes = export_wage_cards_to_excel(cards)
    fname = f"WageCards_{state or 'All'}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    save_audit_entry("DOWNLOAD", filename=fname, details=f"Exported {len(cards)} cards, filter: {state or 'All'}")
    return StreamingResponse(io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.get("/api/config/states")
def list_states():
    return {"states": sorted(set(list(PTAX_SLABS.keys()) + list(LWF_RATES.keys())))}

@app.get("/api/config/statutory")
def get_statutory():
    c = engine.config
    return {"pf_employee_pct": c.pf_employee_pct, "pf_employer_pct": c.pf_employer_pct,
            "pf_employee_cap": c.pf_employee_cap, "pf_employer_cap": c.pf_employer_cap,
            "esic_employee_pct": c.esic_employee_pct, "esic_employer_pct": c.esic_employer_pct,
            "esic_wage_ceiling": c.esic_wage_ceiling, "basic_cap": c.basic_cap,
            "flexi_cap_pct": c.flexi_cap_pct, "hra_cap": c.hra_cap, "nsa_per_night": c.nsa_per_night}

@app.get("/api/config/ptax/{state}")
def get_ptax(state: str):
    slabs = get_ptax_slabs(state)
    return {"state": state, "slabs": [{"threshold": s.threshold, "amount": s.amount} for s in slabs]}

@app.get("/api/config/lwf/{state}")
def get_lwf(state: str):
    l = get_lwf_config(state)
    return {"state": state, "employee_amount": l.employee_amount, "employer_amount": l.employer_amount,
            "frequency": l.frequency, "is_percentage": l.is_percentage}


# --- Upload Excel ---


# --- Associate PT Logic ---

def generate_pt_cards(all_cards, engine, get_ptax_slabs, get_lwf_config):
    """Generate Associate PT cards from Associate cards based on entity-specific rules."""
    import math
    from services.calculation_engine import WageInput
    
    # PT percentage rules
    INFC_SITE_PCT = {'BLR7': 0.50, 'DED3': 0.50, 'BOM5': 0.65, 'PNQ3': 0.65}
    
    # Get all Associate cards
    assoc_cards = [c for c in all_cards if 'associate' in c.get('short_bt', '').lower() 
                   and 'pt' not in c.get('short_bt', '').lower()]
    
    pt_cards = []
    
    for card in assoc_cards:
        entity = card.get('entity', '').upper()
        site = card.get('site_codes', '').upper()
        state = card.get('state', '').upper()
        weekly_hrs = card.get('weekly_hours', 45)
        tenure = card.get('tenure_years', 0)
        
        # Determine PT Gross based on entity logic
        pt_gross = None
        method = None  # 'gross' or 'net'
        
        if entity == 'INFC':
            if site in INFC_SITE_PCT:
                pct = INFC_SITE_PCT[site]
                pt_gross = round(card['gross'] * pct)
                method = 'gross'
            else:
                continue  # Only specific INFC sites have PT
                
        elif entity == 'UFF' or entity == 'AMZL':
            # Net-based: PT Net = 65% of (Associate Net + Default OT if 40hrs)
            assoc_net = card.get('net_salary', 0)
            default_ot = card.get('ot_default', 0)  # Already calculated (OT*22 if 40hrs)
            reference_net = assoc_net + default_ot
            pt_net_target = math.ceil(reference_net * 0.65)  # ROUNDUP
            
            # Reverse calculate Gross from target Net
            # Net = Gross - PF_EE - ESIC_EE
            # PF_EE = MIN(12% * Basic, 1800), ESIC_EE depends on included wages
            # Since Basic = MIN(Gross, 15000), for most PT cases Gross < 15000
            # So Basic = Gross, PF = MIN(12%*Gross, 1800), ESIC = ROUNDUP(Gross*0.75%) if <=21000
            # Net = Gross - MIN(12%*Gross, 1800) - ROUNDUP(Gross*0.75%)
            # Simplified: Net ≈ Gross - 12%*Gross - 0.75%*Gross = Gross * (1 - 0.1275) = Gross * 0.8725
            # Gross ≈ Net / 0.8725
            
            # Iterative approach for accuracy
            is_mh_wb_pt = state in ('MH', 'WB')
            pt_gross = round(pt_net_target / 0.8725)  # Initial estimate
            for _ in range(20):  # Iterate to converge
                # Calculate what net would be at this gross
                if is_mh_wb_pt:
                    test_basic = min(round(pt_gross / 1.05), 15000)
                    test_hra = round(test_basic * 0.05)
                    test_flexi = min(max(0, pt_gross - test_basic - test_hra), 7500)
                else:
                    test_basic = min(pt_gross, 15000)
                    test_flexi = min(max(0, pt_gross - test_basic), 7500)
                test_included = test_basic + test_flexi
                test_pf = min(0.12 * test_basic, 1800)
                test_esic = math.ceil(test_included * 0.0075) if test_included <= 21000 else 0
                test_net = pt_gross - test_pf - test_esic
                
                if abs(test_net - pt_net_target) <= 1:
                    break
                # Adjust
                pt_gross = pt_gross + (pt_net_target - test_net)
                pt_gross = round(pt_gross)
            
            method = 'net'
            
        elif entity == 'ATS':
            if state == 'MH':
                pt_gross = round(card['gross'] * 0.65)
            elif weekly_hrs == 40:
                pt_gross = round(card['gross'] * 0.63)
            else:
                pt_gross = round(card['gross'] * 0.50)
            method = 'gross'
            
        elif entity == 'AMXL':
            if weekly_hrs == 40:
                pt_gross = round(card['gross'] * 0.65)
            else:
                pt_gross = round(card['gross'] * 0.50)
            method = 'gross'
        else:
            continue  # Unknown entity, skip
        
        if pt_gross is None or pt_gross <= 0:
            continue

        # --- IXCE PT GROSS HARDCODE (smart bypass) ---
        # Hard-coded PT Gross values for IXCE. If formula-derived PT Gross exceeds
        # the hardcoded value, bypass the hardcode and use formula instead.
        if site == 'IXCE' and entity == 'AMZL':
            ixce_pt_gross_hardcode = {0: 11071, 1: 11499, 2: 11922, 3: 12133, 4: 12346}
            hardcoded_gross = ixce_pt_gross_hardcode.get(tenure)
            if hardcoded_gross:
                if pt_gross > hardcoded_gross:
                    pass  # Formula exceeds hardcode — use formula (smart bypass)
                else:
                    pt_gross = hardcoded_gross  # Use hardcoded value
        
        # --- PT Attendance Incentive: 50% of matching Associate (same Entity/Site/Year Band) ---
        # The source `card` IS the matching Associate for this Entity + Site + Year Band,
        # so its attendance_incentive is exactly the reference value.
        pt_attendance_incentive = round(card.get('attendance_incentive', 0) * 0.5)

        # Now split and calculate the PT card
        mw = card.get('minimum_wage', 0)
        is_mh_wb = state in ('MH', 'WB')
        
        if is_mh_wb:
            # MH/WB PT: HRA = 5% of Basic, WITHIN Gross (not extra)
            # Basic = Gross / 1.05 (so that Basic + 5%Basic = Gross)
            # Any rounding remainder goes to HRA
            basic_pt = min(round(pt_gross / 1.05), 15000)
            hra_pt = pt_gross - basic_pt  # HRA gets ALL the remainder (no Flexi)
            flexi_pt = 0
            remaining = 0
            lta_pt = 0
            conv_pt = remaining
            split = {"basic": basic_pt, "flexi": flexi_pt, "lta": lta_pt, "hra": hra_pt, "conveyance": conv_pt}
        else:
            split = engine.auto_split_for_mw(
                target_gross=pt_gross,
                minimum_wage=0,  # Ignore MW for PT cards
                state=state,
                hra_applicable=False,
            )
        
        wage_input = WageInput(
            state=card.get('state', ''), city=card.get('city', ''),
            site_code=card.get('site_codes', ''),
            entity=card.get('entity', ''), mw_zone=card.get('mw_zone', ''),
            region=card.get('region', ''), mw_category=card.get('mw_category', ''),
            business_title='Associate PT', short_bt='Associate PT',
            weekly_hours=weekly_hrs,
            daily_hours=card.get('daily_hours', 9),
            minimum_wage=mw, mw_effective_date=card.get('mw_effective_date', ''),
            basic=split['basic'], flexi=split['flexi'], lta=split['lta'],
            hra=split['hra'], conveyance=split['conveyance'],
            attendance_incentive=pt_attendance_incentive, tenure_years=tenure,
        )
        
        ptax = get_ptax_slabs(card.get('state', ''))
        lwf = get_lwf_config(card.get('state', ''))
        result = engine.calculate(wage_input, ptax, lwf)

        # PT has no OT/NSA. Recompute remuneration-dependent values without OT.
        pt_total_remuneration = result.ctc + result.attendance_incentive
        pt_excluded_wages = result.attendance_incentive + result.hra + result.conveyance
        pt_cap_50_amount = round(0.50 * pt_total_remuneration)
        pt_included_pct = round(result.included_wages / pt_total_remuneration, 4) if pt_total_remuneration else 0
        pt_cap_50_met = pt_excluded_wages <= pt_cap_50_amount

        pt_card = {
            "id": f"pt_{card.get('id', '')}",
            "entity": card.get('entity', ''), "state": card.get('state', ''),
            "state_code": card.get('state_code', ''),
            "city": card.get('city', ''), "node": card.get('node', ''),
            "level": card.get('level', ''),
            "mw_zone": card.get('mw_zone', ''), "region": card.get('region', ''),
            "mw_category": card.get('mw_category', ''),
            "business_title": "Associate PT", "short_bt": "Associate PT",
            "site_codes": card.get('site_codes', ''),
            "tenure_years": tenure,
            "weekly_hours": weekly_hrs, "daily_hours": card.get('daily_hours', 9),
            "monthly_ot_limit": None, "minimum_wage": mw,
            "mw_effective_date": card.get('mw_effective_date', ''),
            "basic": result.basic, "flexi": result.flexi, "lta": result.lta,
            "hra": result.hra, "conveyance": result.conveyance, "gross": result.gross,
            "per_hour_ot_total": 0,  # No OT for PT
            "per_hour_ot_included": 0,
            "per_hour_ot_balance": 0,
            "pf_employee": result.pf_employee, "esic_employee": result.esic_employee,
            "pt_employee": "As applicable", "lwf_employee": "As applicable",
            "gross_deductions": result.gross_deductions, "net_salary": result.net_salary,
            "pf_employer": result.pf_employer, "esic_employer": result.esic_employer,
            "lwf_employer": "As applicable", "ctc": result.ctc,
            "ot_default": 0,  # No OT for PT
            "nsa": 0,
            "attendance_incentive": result.attendance_incentive,
            # PT has no OT/NSA: Total Remuneration = CTC + Attendance Incentive
            "total_remuneration": pt_total_remuneration,
            "included_wages": result.included_wages,
            "included_pct": pt_included_pct,
            # PT Excluded = Attendance Incentive + HRA + Conveyance (no OT/NSA)
            "excluded_wages": pt_excluded_wages,
            "cap_50_amount": pt_cap_50_amount,
            "cap_50_met": pt_cap_50_met, "mw_compliant": "N/A",
            "is_pt": True, "pt_method": method,
            "old_ot": 0, "old_hol": 0, "bal_pay_ot": 0, "bal_pay_hol": 0, "hol_wage": 0,
        }
        pt_cards.append(pt_card)
    
    return pt_cards
@app.post("/api/upload-wage-cards")
async def upload_wage_cards(file: UploadFile = File(...), password: str = Form("")):
    """Upload full template. Requires password."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password. Upload access denied.")
    import openpyxl

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.worksheets[0]  # Always use first sheet

    def to_float(val, default=0.0):
        if val is None or val == '' or val == '-':
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def to_str(val, default=''):
        if val is None:
            return default
        return str(val).strip()

    # Read headers from row 1 to find column positions
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column

    # Map expected columns (flexible matching)
    def find_col(names):
        for n in names:
            if n.upper() in headers:
                return headers[n.upper()]
        return None

    col_mile = find_col(['Mile', 'Entity'])
    col_site = find_col(['Site Code', 'Site'])
    col_state = find_col(['State'])
    col_city = find_col(['City'])
    col_node = find_col(['Node', 'Region'])
    col_level = find_col(['Level'])
    col_bt = find_col(['Short BT', 'Business Title'])
    col_mwcat = find_col(['Minimum Wage Category', 'MW Category'])
    col_zone = find_col(['Minimum Wage Zone', 'MW Zone', 'Zone'])
    col_mw = find_col(['Minimum Wage', 'MW', 'MW Amount'])
    col_weekly = find_col(['State Weekly Working Hours', 'Weekly Hours', 'Weekly Hrs'])
    col_daily = find_col(['State Daily Working Hours', 'Daily Hours', 'Daily Hrs'])
    col_0yr = find_col(['0 Year', '0Year', '0 Yr'])
    col_1yr = find_col(['1 Year', '1Year', '1 Yr'])
    col_2yr = find_col(['2 Year', '2Year', '2 Yr'])
    col_3yr = find_col(['3 Year', '3Year', '3 Yr'])
    col_4yr = find_col(['4 Year', '4Year', '4 Yr'])
    col_old_ot = find_col(['Old Per Hr OT', 'Old OT', 'Old Per Hour OT', 'Old OT 0Yr', 'Old OT 0 Yr'])
    col_old_hol = find_col(['Old Hol Wage', 'Old Holiday Wage', 'Old Hol', 'Old Hol 0Yr', 'Old Hol 0 Yr'])
    # Per-year old values (columns after the first Old OT/Hol)
    col_old_ot_1 = find_col(['Old OT 1Yr', 'Old OT 1 Yr'])
    col_old_ot_2 = find_col(['Old OT 2Yr', 'Old OT 2 Yr'])
    col_old_ot_3 = find_col(['Old OT 3Yr', 'Old OT 3 Yr'])
    col_old_ot_4 = find_col(['Old OT 4Yr', 'Old OT 4 Yr'])
    col_old_hol_1 = find_col(['Old Hol 1Yr', 'Old Hol 1 Yr'])
    col_old_hol_2 = find_col(['Old Hol 2Yr', 'Old Hol 2 Yr'])
    col_old_hol_3 = find_col(['Old Hol 3Yr', 'Old Hol 3 Yr'])
    col_old_hol_4 = find_col(['Old Hol 4Yr', 'Old Hol 4 Yr'])

    if not col_state or not col_bt:
        raise HTTPException(400, "Could not find required columns: 'State' and 'Short BT' are mandatory.")
    if not col_0yr:
        raise HTTPException(400, "Could not find Gross columns (0 Year, 1 Year, etc.)")

    # Clear existing cards
    db._memory_store["wage_cards"] = {}

    imported = 0
    errors = []
    now = datetime.utcnow().isoformat()

    for row_idx in range(2, ws.max_row + 1):
        state = to_str(ws.cell(row=row_idx, column=col_state).value) if col_state else ''
        if not state:
            continue

        mile = to_str(ws.cell(row=row_idx, column=col_mile).value) if col_mile else 'AMZL'
        site_code = to_str(ws.cell(row=row_idx, column=col_site).value) if col_site else ''
        city = to_str(ws.cell(row=row_idx, column=col_city).value) if col_city else ''
        node = to_str(ws.cell(row=row_idx, column=col_node).value) if col_node else ''
        level = to_str(ws.cell(row=row_idx, column=col_level).value) if col_level else ''
        short_bt = to_str(ws.cell(row=row_idx, column=col_bt).value) if col_bt else ''
        mw_category = to_str(ws.cell(row=row_idx, column=col_mwcat).value, 'Semi Skilled') if col_mwcat else 'Semi Skilled'
        # Force correct MW Category: Associate/Associate DA = Semi Skilled, all others = Skilled
        if 'associate' in short_bt.lower():
            mw_category = 'Semi Skilled'
        elif short_bt:
            mw_category = 'Skilled'
        mw_zone = to_str(ws.cell(row=row_idx, column=col_zone).value) if col_zone else ''
        minimum_wage = to_float(ws.cell(row=row_idx, column=col_mw).value) if col_mw else 0
        weekly_hours = to_float(ws.cell(row=row_idx, column=col_weekly).value, 45) if col_weekly else 45
        daily_hours = to_float(ws.cell(row=row_idx, column=col_daily).value, 9) if col_daily else 9

        if not short_bt:
            continue

        # Gross rates for each tenure year
        gross_rates = {}
        if col_0yr: gross_rates[0] = to_float(ws.cell(row=row_idx, column=col_0yr).value)
        if col_1yr: gross_rates[1] = to_float(ws.cell(row=row_idx, column=col_1yr).value)
        if col_2yr: gross_rates[2] = to_float(ws.cell(row=row_idx, column=col_2yr).value)
        if col_3yr: gross_rates[3] = to_float(ws.cell(row=row_idx, column=col_3yr).value)
        if col_4yr: gross_rates[4] = to_float(ws.cell(row=row_idx, column=col_4yr).value)

        # Old OT and Hol Wage per tenure year (for balancing pay calculation)
        old_ot_by_year = {
            0: to_float(ws.cell(row=row_idx, column=col_old_ot).value) if col_old_ot else 0,
            1: to_float(ws.cell(row=row_idx, column=col_old_ot_1).value) if col_old_ot_1 else (to_float(ws.cell(row=row_idx, column=col_old_ot).value) if col_old_ot else 0),
            2: to_float(ws.cell(row=row_idx, column=col_old_ot_2).value) if col_old_ot_2 else (to_float(ws.cell(row=row_idx, column=col_old_ot).value) if col_old_ot else 0),
            3: to_float(ws.cell(row=row_idx, column=col_old_ot_3).value) if col_old_ot_3 else (to_float(ws.cell(row=row_idx, column=col_old_ot).value) if col_old_ot else 0),
            4: to_float(ws.cell(row=row_idx, column=col_old_ot_4).value) if col_old_ot_4 else (to_float(ws.cell(row=row_idx, column=col_old_ot).value) if col_old_ot else 0),
        }
        old_hol_by_year = {
            0: to_float(ws.cell(row=row_idx, column=col_old_hol).value) if col_old_hol else 0,
            1: to_float(ws.cell(row=row_idx, column=col_old_hol_1).value) if col_old_hol_1 else (to_float(ws.cell(row=row_idx, column=col_old_hol).value) if col_old_hol else 0),
            2: to_float(ws.cell(row=row_idx, column=col_old_hol_2).value) if col_old_hol_2 else (to_float(ws.cell(row=row_idx, column=col_old_hol).value) if col_old_hol else 0),
            3: to_float(ws.cell(row=row_idx, column=col_old_hol_3).value) if col_old_hol_3 else (to_float(ws.cell(row=row_idx, column=col_old_hol).value) if col_old_hol else 0),
            4: to_float(ws.cell(row=row_idx, column=col_old_hol_4).value) if col_old_hol_4 else (to_float(ws.cell(row=row_idx, column=col_old_hol).value) if col_old_hol else 0),
        }

        # HRA applicability
        hra_applicable = state.upper() in ('MH', 'WB', 'MAHARASHTRA', 'WEST BENGAL')

        # Process each tenure year
        for tenure_yr, gross in gross_rates.items():
            if gross == 0:
                continue

            try:
                split = engine.auto_split_for_mw(
                    target_gross=gross,
                    minimum_wage=minimum_wage,
                    state=state,
                    hra_applicable=hra_applicable,
                )

                wage_input = WageInput(
                    state=state, city=city, site_code=site_code,
                    entity=mile, mw_zone=mw_zone, region=node,
                    mw_category=mw_category, business_title=short_bt,
                    short_bt=short_bt, weekly_hours=weekly_hours,
                    daily_hours=daily_hours, minimum_wage=minimum_wage,
                    mw_effective_date="", basic=split["basic"],
                    flexi=split["flexi"], lta=split["lta"],
                    hra=split["hra"], conveyance=split["conveyance"],
                    attendance_incentive=0, tenure_years=tenure_yr,
                )

                ptax = get_ptax_slabs(state)
                lwf = get_lwf_config(state)
                result = engine.calculate(wage_input, ptax, lwf)

                card = {
                    "id": str(uuid.uuid4()),
                    "entity": mile, "state": state, "state_code": state,
                    "city": city, "node": node, "level": level,
                    "mw_zone": mw_zone or "", "region": node,
                    "mw_category": mw_category, "business_title": short_bt,
                    "short_bt": short_bt, "site_codes": site_code,
                    "tenure_years": tenure_yr,
                    "weekly_hours": weekly_hours, "daily_hours": daily_hours,
                    "monthly_ot_limit": None, "minimum_wage": minimum_wage,
                    "mw_effective_date": "",
                    "basic": result.basic, "flexi": result.flexi, "lta": result.lta,
                    "hra": result.hra, "conveyance": result.conveyance, "gross": result.gross,
                    "per_hour_ot_total": result.per_hour_ot_total,
                    "per_hour_ot_included": result.per_hour_ot_included,
                    "per_hour_ot_balance": result.per_hour_ot_balance,
                    "pf_employee": result.pf_employee, "esic_employee": result.esic_employee,
                    "pt_employee": "As applicable", "lwf_employee": "As applicable",
                    "gross_deductions": result.gross_deductions, "net_salary": result.net_salary,
                    "pf_employer": result.pf_employer, "esic_employer": result.esic_employer,
                    "lwf_employer": "As applicable", "ctc": result.ctc,
                    "ot_default": result.per_hour_ot_total * 22 if weekly_hours == 40 else 0,
                    "nsa": 0, "attendance_incentive": 0,
                    "total_remuneration": result.total_remuneration,
                    "included_wages": result.included_wages,
                    "included_pct": round(result.included_pct, 4),
                    "excluded_wages": result.excluded_wages,
                    "cap_50_amount": result.cap_50_amount,
                    "cap_50_met": result.cap_50_met, "mw_compliant": result.mw_compliant,
                    "old_ot": old_ot_by_year.get(tenure_yr, 0), "old_hol": old_hol_by_year.get(tenure_yr, 0),
                    "bal_pay_ot": max(0, old_ot_by_year.get(tenure_yr, 0) - result.per_hour_ot_total) if old_ot_by_year.get(tenure_yr, 0) > 0 else 0,
                    "bal_pay_hol": max(0, old_hol_by_year.get(tenure_yr, 0) - (round((result.included_wages * 12 / (52 * weekly_hours)) * 1 * daily_hours) if weekly_hours > 0 else 0)) if old_hol_by_year.get(tenure_yr, 0) > 0 else 0,
                    "hol_wage": round((result.included_wages * 12 / (52 * weekly_hours)) * 1 * daily_hours) if weekly_hours > 0 else 0,
                    "created_at": now, "updated_at": now,
                }
                db.put_wage_card(card, skip_save=True)
                imported += 1

            except Exception as e:
                errors.append(f"Row {row_idx}, {tenure_yr}Yr: {str(e)}")

    # --- PARITY ENFORCEMENT ---
    # After all cards are imported, enforce Gross parity across parity groups
    all_for_parity = db.list_wage_cards()
    try:
        parity_updated = enforce_parity(all_for_parity, engine, get_ptax_slabs, get_lwf_config)
        print(f"✅ Parity enforcement: {parity_updated} cards adjusted")
    except Exception as pe:
        print(f"❌ Parity enforcement error: {pe}")
        import traceback
        traceback.print_exc()
        parity_updated = 0

    # Generate Associate PT cards
    all_current = db.list_wage_cards()
    pt_cards = generate_pt_cards(all_current, engine, get_ptax_slabs, get_lwf_config)
    for pc in pt_cards:
        db.put_wage_card(pc, skip_save=True)
    imported += len(pt_cards)

    # Apply AI Depository (Attendance Incentive + Region) to all cards
    ai_dep_applied = apply_ai_depository_to_cards()

    # Save all data once at the end
    db.save()

    # Save uploaded file and log audit trail
    saved_filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    saved_path = os.path.join(UPLOADS_DIR, saved_filename)
    with open(saved_path, 'wb') as f:
        f.write(contents)
    save_audit_entry("UPLOAD", filename=saved_filename, details=f"Imported {imported} cards, {len(errors)} errors, AI applied to {ai_dep_applied}")

    return {
        "status": "completed",
        "mode": "single_tab",
        "imported": imported,
        "parity_adjustments": parity_updated,
        "errors_count": len(errors),
        "errors": errors[:20],
    }

@app.post("/api/enforce-parity")
def api_enforce_parity(password: str = Form("")):
    """Manually trigger parity enforcement on existing data. Requires password."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password.")
    all_cards = db.list_wage_cards()
    parity_updated = enforce_parity(all_cards, engine, get_ptax_slabs, get_lwf_config)
    if parity_updated > 0:
        # Regenerate PT cards after parity adjustment
        all_after = db.list_wage_cards()
        for c in all_after:
            if c.get('is_pt'):
                del db._memory_store["wage_cards"][c["id"]]
        all_non_pt = db.list_wage_cards()
        pt_cards = generate_pt_cards(all_non_pt, engine, get_ptax_slabs, get_lwf_config)
        for pc in pt_cards:
            db.put_wage_card(pc, skip_save=True)
        db.save()
    return {"status": "completed", "parity_adjustments": parity_updated}

@app.get("/api/parity-status")
def get_parity_status():
    """Check current parity status — shows if sites in parity groups have matching components."""
    groups = load_parity_groups()
    all_cards = db.list_wage_cards()
    result = []
    for group in groups:
        group_sites = [s.upper() for s in group.get('sites', [])]
        group_entity = group.get('entity', '').upper() if group.get('entity') else None
        group_cards = [c for c in all_cards 
                       if c.get('site_codes', '').upper() in group_sites 
                       and not c.get('is_pt')
                       and (not group_entity or c.get('entity', '').upper() == group_entity)]
        
        # Check Associate 0Yr
        assoc_0yr = [c for c in group_cards if c.get('short_bt') == 'Associate' and c.get('tenure_years') == 0]
        grosses = list(set(c.get('gross', 0) for c in assoc_0yr))
        ltas = list(set(c.get('lta', 0) for c in assoc_0yr))
        
        sites_detail = []
        for c in sorted(assoc_0yr, key=lambda x: x.get('site_codes', '')):
            sites_detail.append({
                "site": c.get('site_codes', ''),
                "mw_zone": c.get('mw_zone', ''),
                "mw": c.get('minimum_wage', 0),
                "gross": c.get('gross', 0),
                "basic": c.get('basic', 0),
                "flexi": c.get('flexi', 0),
                "lta": c.get('lta', 0),
                "hra": c.get('hra', 0),
                "ot": c.get('per_hour_ot_total', 0),
            })
        
        result.append({
            "group_name": group.get('name', ''),
            "parity_ok": len(grosses) <= 1 and len(ltas) <= 1,
            "unique_grosses": grosses,
            "unique_ltas": ltas,
            "sites": sites_detail,
        })
    
    return {"parity_file_exists": os.path.exists(PARITY_FILE), "parity_file_path": PARITY_FILE, "groups": result}

@app.post("/api/change-password")
def change_password(current_password: str = Form(""), new_password: str = Form("")):
    """Change the upload password."""
    if current_password != get_upload_password():
        raise HTTPException(403, "Current password is incorrect.")
    if len(new_password) < 4:
        raise HTTPException(400, "New password must be at least 4 characters.")
    set_upload_password(new_password)
    return {"status": "success", "message": "Password changed successfully."}

@app.get("/api/audit-trail")
def get_audit_trail():
    """Get all audit trail entries."""
    trail = load_audit_trail()
    trail.reverse()  # Most recent first
    return {"entries": trail}

@app.get("/api/audit-trail/download/{filename}")
def download_audit_file(filename: str):
    """Download a previously uploaded file."""
    filepath = os.path.join(UPLOADS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "File not found")
    return StreamingResponse(
        open(filepath, 'rb'),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# --- Logic Depository ---

LOGIC_TEXT = """
WAGE CARD MANAGEMENT SYSTEM — LOGIC DEPOSITORY
================================================
Developed by: Ravi Kumar (Kmarnuz) | Sr. SME CTK MHLS

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GROSS SPLIT LOGIC — All States (except MH/WB)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Basic       = MIN(Gross, ₹15,000)
2. Flexi       = MIN(remaining after Basic, ₹7,500)
3. LTA         = ONLY if MW > ₹22,500 → LTA = MW - 22,500
4. HRA         = MIN(remaining, ₹7,500)
5. Conveyance  = residual amount

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GROSS SPLIT LOGIC — MH & WB States
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Basic       = MIN(Gross, ₹15,000)
2. HRA         = 5% of MW (mandatory minimum, capped at ₹7,500)
3. Flexi       = MIN(remaining, ₹7,500)
4. LTA         = ONLY if MW > ₹22,500 → LTA = MW - 22,500
5. Conveyance  = residual amount

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WAGE CLASSIFICATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Included Wages (Statutory Base) = Basic + Flexi + LTA
Excluded Wages = HRA + Conveyance + OT + NSA + Incentive

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STATUTORY FORMULAS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PF (Employee)    = MIN(12% × Basic, ₹1,800)
PF (Employer)    = MIN(13% × Basic, ₹1,950)
ESIC (Employee)  = IF(Included > ₹21,000, 0, ROUNDUP(Included × 0.75%))
ESIC (Employer)  = IF(Included > ₹21,000, 0, ROUNDUP(Included × 3.25%))
Per Hour OT      = ROUND((Basic+Flexi+LTA) × 12 / (52 × Weekly Hrs) × 2)
OT Default       = IF(Weekly Hrs = 40, Per Hour OT × 22, 0)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SALARY CALCULATIONS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Gross            = Basic + Flexi + LTA + HRA + Conveyance
Gross Deductions = PF (Employee) + ESIC (Employee)
Net Salary       = Gross - Gross Deductions
CTC              = Gross + PF (Employer) + ESIC (Employer)
Total Remuneration = CTC + Default OT + NSA + Attendance Incentive

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMPLIANCE CHECKS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MW Compliant     = Basic + Flexi + LTA >= Minimum Wage
50% Cap Rule     = Excluded Wages <= 50% of Total Remuneration

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MW ZONE CLASSIFICATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Same city can have different MW for different zones
  (e.g., Bangalore Zone 1, 2, 3)
- MW lookup: State + City + Short BT + Zone + MW Category
- Associate → Semi Skilled
- PA / Supervisor / ADE / Others → Skilled

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMPONENT CAPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Basic            : ₹15,000 maximum
Flexi            : ₹7,500 maximum
HRA              : ₹7,500 maximum (MH/WB: min 5% of MW)
LTA              : Only when MW > ₹22,500 (amount = MW - 22,500)
PF Employee Cap  : ₹1,800/month
PF Employer Cap  : ₹1,950/month
ESIC Ceiling     : ₹21,000 (no ESIC if included wages exceed)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PT & LWF
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Professional Tax  : State-wise slab based (29 states configured)
LWF              : State-wise fixed amounts, varying frequencies
Both shown as "As applicable" in wage card — deducted separately

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ASSOCIATE PT WAGE CARD LOGIC
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Auto-derived from Associate regular cards. OT = 0. MW = N/A.

Entity Rules:
  INFC BLR7/DED3  : PT Gross = 50% of Associate Gross
  INFC BOM5/PNQ3  : PT Gross = 65% of Associate Gross
  AMZL (All)      : PT Net = ROUNDUP(65% x (Assoc Net + Default OT))
  UFF (All)       : PT Net = ROUNDUP(65% x (Assoc Net + Default OT))
  ATS MH          : PT Gross = 65% of Associate Gross
  ATS 40hr sites  : PT Gross = 63% of Associate Gross
  ATS others      : PT Gross = 50% of Associate Gross
  AMXL 40hr sites : PT Gross = 65% of Associate Gross
  AMXL others     : PT Gross = 50% of Associate Gross

MH/WB PT Split:
  Basic = PT Gross / 1.05 (so HRA fits within Gross)
  HRA = 5% of Basic
  Basic + HRA = PT Gross exactly

MW Category:
  Associate & Associate PT = Semi Skilled
  All other roles = Skilled

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TEMPLATE FORMAT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Single Tab "Wage Card Depository":
  Mile | Site Code | State | City | Weekly Hrs | Daily Hrs | Node | Level |
  Short BT | MW Category | MW Zone | Minimum Wage | MW Effective Date |
  0 Year | 1 Year | 2 Year | 3 Year | 4 Year |
  Old OT 0-4Yr | Old Hol 0-4Yr

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MW REVISION LOGIC
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
When new MW is uploaded via MW Revision:
  If New MW <= 0Yr Gross : Keep Gross unchanged, only re-split components
  If New MW > 0Yr Gross  : Gap = New MW - 0Yr Gross
    All tenure years increased by Gap (0Yr = MW, 1-4Yr = old + Gap)
    Then re-split and recalculate all components

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PARITY GROUPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Sites in a parity group maintain IDENTICAL Gross and component split
(Basic, Flexi, LTA, HRA, Conveyance, OT) regardless of MW zone differences.

After any upload or MW revision, the HIGHEST Gross in the group is applied
to ALL sites, with the exact same component split copied from the reference.

Configured Parity Groups:
  1. UFF Bangalore: SBLY, SBLZ, UBL5, UBL6, UBL9
     - SBLZ is Zone 2 but follows Zone 1 rates
  2. UFF Delhi-NCR: PDL1, PDL2, PDL5, UDL4
     - HY sites (MW 16,781) match DL sites (MW 20,371+)
  3. AMZL Delhi-NCR-Faridabad: DELH, DLIH, FADA, NCT2, NZMN,
     DELF, DELG, DELK, DELL, DELN, DELO, DELR, DELT,
     NCRG, NCTC, NCTD, NCTG, NDBA, NZML
     - HY (MW 18,501) and DL (MW 22,411) share same Gross

Config file: parity_groups.json

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
IXCE ASSOCIATE PT HARDCODE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
IXCE (AMZL, Panchkula) has hard-coded Associate PT Gross values:
  0 Year: 11,071
  1 Year: 11,499
  2 Year: 11,922
  3 Year: 12,133
  4 Year: 12,346

Smart Bypass: If formula-derived PT Gross exceeds the hardcoded value
(after a future MW revision), the hardcode is automatically dropped.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ALFA RATE CARD LOGIC
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Premium: 15% applied to all components (Basic, Flexi, LTA, HRA)
P10 = Premium components / 22 (daily rate for 10 days)
P5  = Premium components / 22 / Daily Hrs * 4
P8  = Premium components / 22 / Daily Hrs * 8
OT  = ROUND(((Basic_P+Flexi_P+LTA_P)*12/(52*WeeklyHrs))*2) — only for 40hr sites

Hard-Coded P10 OT (config: alfa_hardcode.json):
  - UDL6 (UFF, Noida): P10_OT = 227
  - LKOI, IXDD, KNUD, KNUO, AGRD, GKPL, MREE, VNSD, LKOA, LKOD (AMZL): P10_OT = 211
  - LKO1 (INFC, Lucknow): P10_OT = 211
  - LKOO (ATS, Lucknow): P10_OT = 211
  - NCRJ, NCT3, NCT8, NZMF, NZMM (AMZL, Ghaziabad/Noida): P10_OT = 227
  Smart Bypass: If formula OT > hardcoded OT after rate increase, formula used.
  Net Pay flows naturally: Net = (P10_Gross + P10_OT) - Deductions

ROUND PF Sites (use ROUND instead of ROUNDUP for PF calculation):
  - FHYE (UFF): P10=1001, P5=386, P8=775
  - SBLZ, UBL6, UBL9, UBL5, SBLY (UFF): P10=1291

Holiday Pay Override (INFC MH/GJ + HMH4):
  Standard:  PH10 = ((Basic_P+Flexi_P+LTA_P)*12/(52*WeeklyHrs))*DailyHrs
  Override:  PH10 = ((Basic_P+Flexi_P+LTA_P)*12)/(52*45)*2*9
             PH5  = ((Basic_P+Flexi_P+LTA_P)*12)/(52*45)*2*4
             PH8  = PH5 * 2

Smart Bypass for ALFA Hardcodes:
  If Gross increases and formula P10_OT > hardcoded P10_OT,
  the hardcode is automatically bypassed and formula is used.
  Flagged in "Remarks & Reference" sheet of ALFA export.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BALANCING PAY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Old Per Hr OT and Old Hol Wage: From template (reference values)
Bal Pay OT   = MAX(0, Old OT - New Per Hr OT)
Bal Pay Hol  = MAX(0, Old Hol - New Hol Wage)
PT cards: Blank (not applicable)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ROUNDING RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
All monetary values rounded to whole numbers (no decimals).
PF: ROUNDUP (except ROUND PF sites listed above)
ESIC: ROUNDUP
OT/Hol Wage: ROUND
"""

@app.get("/api/logic-depository")
def get_logic_depository():
    """Get the logic depository content."""
    return {"content": LOGIC_TEXT}

@app.get("/api/logic-depository/download")
def download_logic_depository():
    """Download the logic depository as a text file."""
    return StreamingResponse(
        io.BytesIO(LOGIC_TEXT.encode('utf-8')),
        media_type="text/plain",
        headers={"Content-Disposition": "attachment; filename=Wage_Card_Logic_Depository.txt"}
    )


# ============================================================
# FRONTEND — Served from the same app
# ============================================================



@app.get("/api/revision-template/download")
def download_revision_template():
    """Download a blank MW Revision template with headers and data validation."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MW Revision"

    # Headers
    headers = ["Site Code", "Short BT", "Minimum Wage Category", "Minimum Wage Zone", "Minimum Wage", "Effective Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # Get existing data for dropdowns
    all_cards = db.list_wage_cards()
    sites = sorted(set(c.get("site_codes", "") for c in all_cards if c.get("site_codes")))
    bts = sorted(set(c.get("short_bt", "") for c in all_cards if c.get("short_bt")))

    # Add data validation for Site Code (column A)
    if sites:
        site_list = ",".join(sites[:200])  # Excel limit
        dv_site = DataValidation(type="list", formula1=f'"{site_list}"', allow_blank=False)
        dv_site.error = "Please select a valid Site Code"
        dv_site.errorTitle = "Invalid Site"
        ws.add_data_validation(dv_site)
        dv_site.add(f"A2:A1000")

    # Add data validation for Short BT (column B)
    if bts:
        bt_list = ",".join(bts[:50])
        dv_bt = DataValidation(type="list", formula1=f'"{bt_list}"', allow_blank=False)
        dv_bt.error = "Please select a valid Short BT"
        dv_bt.errorTitle = "Invalid BT"
        ws.add_data_validation(dv_bt)
        dv_bt.add(f"B2:B1000")

    # Add data validation for MW Category (column C)
    dv_cat = DataValidation(type="list", formula1='"Semi Skilled,Skilled"', allow_blank=False)
    ws.add_data_validation(dv_cat)
    dv_cat.add(f"C2:C1000")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Add instruction sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        "MW REVISION TEMPLATE — Instructions",
        "",
        "Fill in this template to update Minimum Wage for specific sites.",
        "",
        "MANDATORY columns:",
        "  - Site Code: Select from dropdown (existing sites only)",
        "  - Short BT: Select from dropdown (Associate, PA, etc.)",
        "  - Minimum Wage Category: Semi Skilled or Skilled",
        "  - Minimum Wage: New MW amount",
        "  - Effective Date: Date format YYYY-MM-DD",
        "",
        "OPTIONAL columns:",
        "  - Minimum Wage Zone: Only if zone classification applies",
        "",
        "HOW IT WORKS:",
        "  - If new MW <= existing 0Yr Gross: Gross unchanged, only re-split for compliance",
        "  - If new MW > existing 0Yr Gross: All year Gross increased by the gap",
        "  - All statutory fields recalculated automatically",
        "",
        "Upload via the '🔄 MW Revision' button on the dashboard.",
    ]
    for i, line in enumerate(instructions, 1):
        ws_info.cell(row=i, column=1, value=line)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=MW_Revision_Template.xlsx"}
    )

@app.post("/api/upload-revision")
async def upload_revision(file: UploadFile = File(...), password: str = Form("")):
    """Upload MW Revision. Requires password."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password. Upload access denied.")
    import openpyxl

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.worksheets[0]

    def to_float(val, default=0.0):
        if val is None or val == '' or val == '-':
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def to_str(val, default=''):
        if val is None:
            return default
        return str(val).strip()

    # Read headers
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column

    def find_col(names):
        for n in names:
            if n.upper() in headers:
                return headers[n.upper()]
        return None

    col_site = find_col(['Site Code', 'Site'])
    col_bt = find_col(['Short BT', 'Business Title'])
    col_mwcat = find_col(['Minimum Wage Category', 'MW Category'])
    col_zone = find_col(['Minimum Wage Zone', 'MW Zone', 'Zone'])
    col_mw = find_col(['Minimum Wage', 'MW', 'New Minimum Wage'])
    col_date = find_col(['Effective Date', 'MW Effective Date', 'Minimum Wage (Effective Date)'])

    if not col_site:
        raise HTTPException(400, "Required column 'Site Code' not found.")
    if not col_mw:
        raise HTTPException(400, "Required column 'Minimum Wage' not found.")

    # Build revision lookup: site_code + short_bt -> new MW data
    revisions = {}
    for row_idx in range(2, ws.max_row + 1):
        site = to_str(ws.cell(row=row_idx, column=col_site).value)
        if not site:
            continue
        bt = to_str(ws.cell(row=row_idx, column=col_bt).value) if col_bt else ''
        mw_cat = to_str(ws.cell(row=row_idx, column=col_mwcat).value) if col_mwcat else ''
        zone = to_str(ws.cell(row=row_idx, column=col_zone).value) if col_zone else ''
        new_mw = to_float(ws.cell(row=row_idx, column=col_mw).value)
        eff_date = ''
        if col_date:
            d = ws.cell(row=row_idx, column=col_date).value
            if hasattr(d, 'strftime'):
                eff_date = d.strftime('%Y-%m-%d')
            else:
                eff_date = to_str(d)

        if new_mw > 0:
            key = (site.upper(), bt.upper() if bt else None)
            revisions[key] = {
                "mw": new_mw,
                "mw_category": mw_cat,
                "zone": zone,
                "effective_date": eff_date,
            }

    # Apply revisions to existing cards
    all_cards = db.list_wage_cards()
    updated = 0
    now = datetime.utcnow().isoformat()

    # Group cards by site+bt to find 0Yr gross for gap calculation
    from collections import defaultdict
    card_groups = defaultdict(list)
    for card in all_cards:
        key = (card.get("site_codes", "").upper(), card.get("short_bt", "").upper())
        card_groups[key].append(card)

    for (site_code, short_bt), cards_in_group in card_groups.items():
        # Find matching revision
        rev = revisions.get((site_code, short_bt))
        if not rev:
            rev = revisions.get((site_code, None))
        if not rev:
            for k, v in revisions.items():
                if k[0] == site_code and (k[1] is None or k[1] in short_bt or short_bt in k[1]):
                    rev = v
                    break
        if not rev:
            continue

        new_mw = rev["mw"]

        # Find 0 Year card to determine gap
        yr0_card = next((c for c in cards_in_group if c.get("tenure_years") == 0), None)
        if not yr0_card:
            continue

        yr0_gross = yr0_card.get("gross", 0)

        # Determine if Gross needs to increase
        if new_mw > yr0_gross:
            # MW > 0Yr Gross: increase all years by the gap
            gap = new_mw - yr0_gross
        else:
            # MW <= 0Yr Gross: just re-split, no Gross change
            gap = 0

        # Apply to all tenure cards in this group
        for card in cards_in_group:
            card["minimum_wage"] = new_mw
            if rev["effective_date"]:
                card["mw_effective_date"] = rev["effective_date"]
            if rev["zone"]:
                card["mw_zone"] = rev["zone"]
            if rev["mw_category"]:
                card["mw_category"] = rev["mw_category"]

            # Adjust Gross if needed
            new_gross = card["gross"] + gap

            # Re-split with new MW and new Gross
            hra_applicable = card.get("state", "").upper() in ('MH', 'WB')
            split = engine.auto_split_for_mw(
                target_gross=new_gross,
                minimum_wage=new_mw,
                state=card.get("state", ""),
                hra_applicable=hra_applicable,
            )

            # Recalculate everything
            wage_input = WageInput(
                state=card.get("state", ""), city=card.get("city", ""),
                site_code=card.get("site_codes", ""),
                entity=card.get("entity", ""), mw_zone=card.get("mw_zone", ""),
                region=card.get("region", ""), mw_category=card.get("mw_category", ""),
                business_title=card.get("short_bt", ""), short_bt=card.get("short_bt", ""),
                weekly_hours=card.get("weekly_hours", 45),
                daily_hours=card.get("daily_hours", 9),
                minimum_wage=new_mw, mw_effective_date=rev.get("effective_date", ""),
                basic=split["basic"], flexi=split["flexi"], lta=split["lta"],
                hra=split["hra"], conveyance=split["conveyance"],
                attendance_incentive=card.get("attendance_incentive", 0),
                tenure_years=card.get("tenure_years", 0),
            )

            ptax = get_ptax_slabs(card.get("state", ""))
            lwf = get_lwf_config(card.get("state", ""))
            result = engine.calculate(wage_input, ptax, lwf)

            # Update card fields
            card.update({
                "basic": result.basic, "flexi": result.flexi, "lta": result.lta,
                "hra": result.hra, "conveyance": result.conveyance, "gross": result.gross,
                "per_hour_ot_total": result.per_hour_ot_total,
                "per_hour_ot_included": result.per_hour_ot_included,
                "per_hour_ot_balance": result.per_hour_ot_balance,
                "pf_employee": result.pf_employee, "esic_employee": result.esic_employee,
                "gross_deductions": result.gross_deductions, "net_salary": result.net_salary,
                "pf_employer": result.pf_employer, "esic_employer": result.esic_employer,
                "ctc": result.ctc,
                "ot_default": result.per_hour_ot_total * 22 if card.get("weekly_hours", 45) == 40 else 0,
                "total_remuneration": result.total_remuneration,
                "included_wages": result.included_wages,
                "included_pct": round(result.included_pct, 4),
                "excluded_wages": result.excluded_wages,
                "cap_50_amount": result.cap_50_amount,
                "cap_50_met": result.cap_50_met, "mw_compliant": result.mw_compliant,
                "updated_at": now,
            })

            db.put_wage_card(card, skip_save=True)
            updated += 1

    # --- PARITY ENFORCEMENT ---
    # After individual MW revisions, enforce Gross parity across parity groups
    all_cards_for_parity = db.list_wage_cards()
    try:
        parity_updated = enforce_parity(all_cards_for_parity, engine, get_ptax_slabs, get_lwf_config)
        print(f"✅ MW Revision parity enforcement: {parity_updated} cards adjusted")
    except Exception as pe:
        print(f"❌ MW Revision parity error: {pe}")
        import traceback
        traceback.print_exc()
        parity_updated = 0

    # Regenerate PT cards after MW revision
    # Remove old PT cards first
    all_after = db.list_wage_cards()
    for c in all_after:
        if c.get('is_pt'):
            del db._memory_store["wage_cards"][c["id"]]
    # Generate new PT cards
    all_non_pt = db.list_wage_cards()
    pt_cards = generate_pt_cards(all_non_pt, engine, get_ptax_slabs, get_lwf_config)
    for pc in pt_cards:
        db.put_wage_card(pc, skip_save=True)

    # Apply AI Depository (Attendance Incentive + Region) after MW revision
    apply_ai_depository_to_cards()

    db.save()

    # Audit
    saved_filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_revision_{file.filename}"
    saved_path = os.path.join(UPLOADS_DIR, saved_filename)
    with open(saved_path, 'wb') as f:
        f.write(contents)
    save_audit_entry("MW_REVISION", filename=saved_filename, details=f"Updated {updated} cards from {len(revisions)} revision entries, {parity_updated} parity adjustments")

    return {
        "status": "completed",
        "mode": "mw_revision",
        "revision_entries": len(revisions),
        "cards_updated": updated,
        "parity_adjustments": parity_updated,
    }



# --- Export Current Template (Updated Dump) ---

@app.get("/api/export-template")
def export_current_template():
    """Export current app data back into the upload template format.
    Use this to get an updated template after MW revisions."""
    import openpyxl
    from openpyxl.styles import Font, Border, Side

    all_cards = db.list_wage_cards()
    if not all_cards:
        raise HTTPException(404, "No data to export.")

    # Group cards by site+bt (to get all tenure years in one row)
    from collections import defaultdict
    groups = defaultdict(dict)
    for card in all_cards:
        key = (card.get('entity',''), card.get('site_codes',''), card.get('state',''),
               card.get('city',''), card.get('region',''), card.get('level',''),
               card.get('short_bt',''), card.get('mw_category',''), card.get('mw_zone',''),
               card.get('minimum_wage',0), card.get('weekly_hours',45), card.get('daily_hours',9))
        tenure = card.get('tenure_years', 0)
        groups[key][tenure] = card.get('gross', 0)
        groups[key][f'old_ot_{tenure}'] = card.get('old_ot', 0)
        groups[key][f'old_hol_{tenure}'] = card.get('old_hol', 0)
        # Store other fields from any card in the group
        if 'card_ref' not in groups[key]:
            groups[key]['card_ref'] = card

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wage Card Depository"

    # Headers matching Template_V2 format
    headers = ["Mile", "Site Code", "State", "City", "State Weekly Working Hours",
               "State Daily Working Hours", "Node", "Level", "Short BT",
               "Minimum Wage Category", "Minimum Wage Zone", "Minimum Wage",
               "Minimum Wage (Effective Date)", "0 Year", "1 Year", "2 Year", "3 Year", "4 Year",
               "Old OT 0Yr", "Old OT 1Yr", "Old OT 2Yr", "Old OT 3Yr", "Old OT 4Yr",
               "Old Hol 0Yr", "Old Hol 1Yr", "Old Hol 2Yr", "Old Hol 3Yr", "Old Hol 4Yr"]

    hdr_font = Font(bold=True, size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.border = border

    # Write data rows
    row_idx = 2
    for key, tenure_data in sorted(groups.items(), key=lambda x: (x[0][2], x[0][3], x[0][6])):
        card = tenure_data.get('card_ref', {})
        entity, site, state, city, node, level, bt, mw_cat, mw_zone, mw, weekly, daily = key

        ws.cell(row=row_idx, column=1, value=entity)
        ws.cell(row=row_idx, column=2, value=site)
        ws.cell(row=row_idx, column=3, value=state)
        ws.cell(row=row_idx, column=4, value=city)
        ws.cell(row=row_idx, column=5, value=weekly)
        ws.cell(row=row_idx, column=6, value=daily)
        ws.cell(row=row_idx, column=7, value=node)
        ws.cell(row=row_idx, column=8, value=level)
        ws.cell(row=row_idx, column=9, value=bt)
        ws.cell(row=row_idx, column=10, value=mw_cat)
        ws.cell(row=row_idx, column=11, value=mw_zone if mw_zone else None)
        ws.cell(row=row_idx, column=12, value=mw)
        ws.cell(row=row_idx, column=13, value=card.get('mw_effective_date', ''))
        ws.cell(row=row_idx, column=14, value=tenure_data.get(0, 0))
        ws.cell(row=row_idx, column=15, value=tenure_data.get(1, 0))
        ws.cell(row=row_idx, column=16, value=tenure_data.get(2, 0))
        ws.cell(row=row_idx, column=17, value=tenure_data.get(3, 0))
        ws.cell(row=row_idx, column=18, value=tenure_data.get(4, 0))

        # Old OT 0-4 Year (columns 19-23)
        ws.cell(row=row_idx, column=19, value=round(tenure_data.get('old_ot_0', 0) or 0))
        ws.cell(row=row_idx, column=20, value=round(tenure_data.get('old_ot_1', 0) or 0))
        ws.cell(row=row_idx, column=21, value=round(tenure_data.get('old_ot_2', 0) or 0))
        ws.cell(row=row_idx, column=22, value=round(tenure_data.get('old_ot_3', 0) or 0))
        ws.cell(row=row_idx, column=23, value=round(tenure_data.get('old_ot_4', 0) or 0))

        # Old Hol 0-4 Year (columns 24-28)
        ws.cell(row=row_idx, column=24, value=round(tenure_data.get('old_hol_0', 0) or 0))
        ws.cell(row=row_idx, column=25, value=round(tenure_data.get('old_hol_1', 0) or 0))
        ws.cell(row=row_idx, column=26, value=round(tenure_data.get('old_hol_2', 0) or 0))
        ws.cell(row=row_idx, column=27, value=round(tenure_data.get('old_hol_3', 0) or 0))
        ws.cell(row=row_idx, column=28, value=round(tenure_data.get('old_hol_4', 0) or 0))

        for col in range(1, 29):
            ws.cell(row=row_idx, column=col).border = border

        row_idx += 1

    # Column widths
    for col in range(1, 29):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Template_Updated_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    save_audit_entry("DOWNLOAD", filename=fname, details=f"Updated template exported ({row_idx-2} rows)")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# --- P-TAX Depository ---

@app.get("/api/ptax-depository")
def get_ptax_depository():
    """Get complete P-TAX depository for all states."""
    from config.statutory_data import PTAX_SLABS, STATE_CODE_MAP
    result = {}
    for state, slabs in PTAX_SLABS.items():
        # Find state code
        code = next((k for k, v in STATE_CODE_MAP.items() if v == state), state)
        result[state] = {
            "code": code,
            "slabs": [{"threshold": s.threshold, "amount": s.amount,
                      "special_month": s.special_month, "special_amount": s.special_amount}
                     for s in slabs]
        }
    return {"states": result, "total_states": len(result)}

@app.get("/api/ptax-depository/download")
def download_ptax_depository():
    """Download P-TAX depository as Excel."""
    import openpyxl
    from openpyxl.styles import Font, Border, Side
    from config.statutory_data import PTAX_SLABS, STATE_CODE_MAP

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PTAX Depository"

    headers = ["State", "State Code", "Gross Earnings Threshold", "PTAX Amount", "Remarks"]
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.border = border

    row_idx = 2
    for state in sorted(PTAX_SLABS.keys()):
        code = next((k for k, v in STATE_CODE_MAP.items() if v == state), "")
        slabs = PTAX_SLABS[state]
        for slab in slabs:
            ws.cell(row=row_idx, column=1, value=state).border = border
            ws.cell(row=row_idx, column=2, value=code).border = border
            ws.cell(row=row_idx, column=3, value=slab.threshold).border = border
            ws.cell(row=row_idx, column=4, value=slab.amount).border = border
            remark = f"{slab.special_amount} in {slab.special_month}" if slab.special_month else ""
            ws.cell(row=row_idx, column=5, value=remark).border = border
            row_idx += 1

    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=PTAX_Depository_2026.xlsx"})

@app.post("/api/ptax-depository/upload")
async def upload_ptax_depository(file: UploadFile = File(...), password: str = Form("")):
    """Upload updated P-TAX depository to replace existing slabs."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password.")

    from config.statutory_data import PTAX_SLABS, PTaxSlab
    import openpyxl

    contents = await file.read()

    # Try .xlsb format
    try:
        from pyxlsb import open_workbook as open_xlsb
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsb')
        tmp.write(contents)
        tmp.close()
        
        new_slabs = {}
        with open_xlsb(tmp.name) as wbx:
            sheet_name = wbx.sheets[0]
            with wbx.get_sheet(sheet_name) as sheet:
                first = True
                for row in sheet.rows():
                    if first:
                        first = False
                        continue
                    vals = [cell.v for cell in row]
                    if not vals or not vals[0]:
                        continue
                    state = str(vals[0]).strip().upper()
                    threshold = float(vals[1]) if vals[1] else 0
                    amount = float(vals[2]) if vals[2] else 0
                    if state not in new_slabs:
                        new_slabs[state] = []
                    new_slabs[state].append(PTaxSlab(threshold=threshold, amount=amount))
        os.unlink(tmp.name)
    except:
        # Try .xlsx format
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.worksheets[0]
        new_slabs = {}
        for row_idx in range(2, ws.max_row + 1):
            state = str(ws.cell(row=row_idx, column=1).value or '').strip().upper()
            threshold = float(ws.cell(row=row_idx, column=3).value or 0)
            amount = float(ws.cell(row=row_idx, column=4).value or 0)
            if state:
                if state not in new_slabs:
                    new_slabs[state] = []
                new_slabs[state].append(PTaxSlab(threshold=threshold, amount=amount))

    if not new_slabs:
        raise HTTPException(400, "No valid PTAX data found in file.")

    # Update the global PTAX_SLABS
    PTAX_SLABS.clear()
    PTAX_SLABS.update(new_slabs)

    save_audit_entry("PTAX_UPDATE", filename=file.filename, details=f"Updated {len(new_slabs)} states")

    return {"status": "success", "states_updated": len(new_slabs)}
# --- ALFA Rate Card Export ---

@app.get("/api/alfa-rate-card/export")
def export_alfa_rate_card():
    """Generate ALFA Rate Card with Excel formulas from Associate 0-Year rates."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import math

    # Colors for ALFA
    FILL_NET = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")  # Mid green - Net Pay
    FILL_GROSS = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    FILL_COMP = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    # Net columns: 32(P10_Net), 43(P5_Net), 54(P8_Net), 56(PH10_Net), 58(PH5_Net), 60(PH8_Net)
    # Gross columns: 24(P10_Gross), 37(P5_Gross), 48(P8_Gross), 19(Gross_P)
    # Component columns: 15(Basic_P), 16(Flexi_P), 17(LTA_P)
    NET_COLS = {32, 43, 54, 56, 58, 60}
    GROSS_COLS = {19, 24, 26, 37, 48}
    COMP_COLS = {15, 16, 17}

    # Load PT/LWF reference
    alfa_pt_lwf_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'alfa_pt_lwf.json')
    alfa_pt_lwf = {}
    if os.path.exists(alfa_pt_lwf_file):
        with open(alfa_pt_lwf_file, 'r') as f:
            alfa_pt_lwf = json.load(f)

    # Get Associate 0-Year cards
    all_cards = db.list_wage_cards()
    assoc_cards = [c for c in all_cards if c.get('tenure_years') == 0 and c.get('short_bt', '') == 'Associate']

    if not assoc_cards:
        raise HTTPException(404, "No Associate 0-Year cards found. Upload main template first.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALFA Rate Card"

    # Column layout (for formula references)
    # A=SlNo B=State C=City D=Site E=DailyHrs F=WeeklyHrs G=MandOT H=MW
    # I=Basic J=Flexi K=LTA L=HRA M=Gross N=Premium%
    # O=Basic_P P=Flexi_P Q=LTA_P R=HRA_P S=Gross_P
    # T=P10_Basic U=P10_Flexi V=P10_LTA W=P10_HRA X=P10_Gross
    # Y=P10_OT Z=P10_TotalGross AA=P10_PF AB=P10_ESIC AC=P10_PT AD=P10_LWF AE=P10_TotDed AF=P10_Net
    # AG=P5_Basic AH=P5_Flexi AI=P5_LTA AJ=P5_HRA AK=P5_Gross
    # AL=P5_PF AM=P5_ESIC AN=P5_PT AO=P5_LWF AP=P5_TotDed AQ=P5_Net
    # AR=P8_Basic AS=P8_Flexi AT=P8_LTA AU=P8_HRA AV=P8_Gross
    # AW=P8_PF AX=P8_ESIC AY=P8_PT AZ=P8_LWF BA=P8_TotDed BB=P8_Net
    # BC=PH10_ESIC BD=PH10_Net BE=PH5_ESIC BF=PH5_Net BG=PH8_ESIC BH=PH8_Net
    # BI=PH10 BJ=PH5 BK=PH8

    headers = [
        "Entity", "State", "City", "Site Code", "Daily Hrs", "Weekly Hrs",
        "Mandatory OT", "Min Wage", "Basic", "Flexi", "LTA", "HRA", "Gross",
        "Premium %", "Basic_P", "Flexi_P", "LTA_P", "HRA_P", "Gross_P",
        "P10_Basic", "P10_Flexi", "P10_LTA", "P10_HRA", "P10_Gross",
        "P10_OT", "P10_Total Gross", "P10_PF", "P10_ESIC", "P10_PT", "P10_LWF",
        "P10_Tot Ded", "P10_Net",
        "P5_Basic", "P5_Flexi", "P5_LTA", "P5_HRA", "P5_Gross",
        "P5_PF", "P5_ESIC", "P5_PT", "P5_LWF", "P5_Tot Ded", "P5_Net",
        "P8_Basic", "P8_Flexi", "P8_LTA", "P8_HRA", "P8_Gross",
        "P8_PF", "P8_ESIC", "P8_PT", "P8_LWF", "P8_Tot Ded", "P8_Net",
        "PH10_ESIC", "PH10_Net", "PH5_ESIC", "PH5_Net", "PH8_ESIC", "PH8_Net",
        "PH10", "PH5", "PH8",
    ]

    hf = Font(bold=True, size=9)
    hdr_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
    hdr_font = Font(bold=True, size=9, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = THIN_BORDER

    # Load ALFA hardcode overrides
    alfa_hardcode_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'alfa_hardcode.json')
    alfa_hardcode = {}
    if os.path.exists(alfa_hardcode_file):
        with open(alfa_hardcode_file, 'r') as f:
            alfa_hardcode = json.load(f)

    # Build lookup for hardcoded P10 OT: (site_code, attendance_code) -> ot_amount
    hardcoded_p10_ot = {}
    for entry in alfa_hardcode.get('hardcoded_p10_ot', []):
        key = (entry['site_code'].upper(), entry['attendance_code'].upper())
        hardcoded_p10_ot[key] = entry['ot_amount']

    # Build lookup for ROUND PF sites: site_code -> attendance_codes list
    round_pf_sites = {}
    for entry in alfa_hardcode.get('round_pf_sites', []):
        round_pf_sites[entry['site_code'].upper()] = [c.upper() for c in entry.get('attendance_codes', [])]

    # Track hardcode bypass notifications
    hardcode_notifications = []

    def compute_p10_net(basic, flexi, lta, hra, weekly_hours, use_round, pt_p10, lwf_p10):
        """Compute P10 Net Pay using formula logic (Python-side) for comparison."""
        premium = 1.15
        basic_p = round(basic * premium)
        flexi_p = round(flexi * premium)
        lta_p = round(lta * premium)
        hra_p = round(hra * premium)
        p10_basic = round(basic_p / 22)
        p10_flexi = round(flexi_p / 22)
        p10_lta = round(lta_p / 22)
        p10_hra = round(hra_p / 22)
        p10_gross = p10_basic + p10_flexi + p10_lta + p10_hra
        included_p = basic_p + flexi_p + lta_p
        if weekly_hours == 40:
            p10_ot = round(((included_p) * 12 / (52 * weekly_hours)) * 2)
        else:
            p10_ot = 0
        p10_total = p10_gross + p10_ot
        if use_round:
            p10_pf = round(p10_basic * 0.12)
        else:
            p10_pf = math.ceil(p10_basic * 0.12)
        if included_p > 21000:
            p10_esic = 0
        else:
            p10_esic = math.ceil((p10_basic + p10_flexi + p10_lta) * 0.0075)
        p10_tot_ded = p10_pf + p10_esic + pt_p10 + lwf_p10
        p10_net = p10_total - p10_tot_ded
        return p10_net

    def compute_p5_net(basic, flexi, lta, hra, daily_hours, weekly_hours, use_round, pt_p5, lwf_p5):
        """Compute P5 Net Pay."""
        premium = 1.15
        basic_p = round(basic * premium)
        flexi_p = round(flexi * premium)
        lta_p = round(lta * premium)
        hra_p = round(hra * premium)
        p5_basic = round(basic_p / 22 / daily_hours * 4)
        p5_flexi = round(flexi_p / 22 / daily_hours * 4)
        p5_lta = round(lta_p / 22 / daily_hours * 4)
        p5_hra = round(hra_p / 22 / daily_hours * 4)
        p5_gross = p5_basic + p5_flexi + p5_lta + p5_hra
        included_p = basic_p + flexi_p + lta_p
        if use_round:
            p5_pf = round(p5_basic * 0.12)
        else:
            p5_pf = math.ceil(p5_basic * 0.12)
        if included_p > 21000:
            p5_esic = 0
        else:
            p5_esic = math.ceil((p5_basic + p5_flexi + p5_lta) * 0.0075)
        p5_tot_ded = p5_pf + p5_esic + pt_p5 + lwf_p5
        p5_net = p5_gross - p5_tot_ded
        return p5_net

    def compute_p8_net(basic, flexi, lta, hra, daily_hours, weekly_hours, use_round, pt_p8, lwf_p8):
        """Compute P8 Net Pay."""
        premium = 1.15
        basic_p = round(basic * premium)
        flexi_p = round(flexi * premium)
        lta_p = round(lta * premium)
        hra_p = round(hra * premium)
        p8_basic = round(basic_p / 22 / daily_hours * 8)
        p8_flexi = round(flexi_p / 22 / daily_hours * 8)
        p8_lta = round(lta_p / 22 / daily_hours * 8)
        p8_hra = round(hra_p / 22 / daily_hours * 8)
        p8_gross = p8_basic + p8_flexi + p8_lta + p8_hra
        included_p = basic_p + flexi_p + lta_p
        if use_round:
            p8_pf = round(p8_basic * 0.12)
        else:
            p8_pf = math.ceil(p8_basic * 0.12)
        if included_p > 21000:
            p8_esic = 0
        else:
            p8_esic = math.ceil((p8_basic + p8_flexi + p8_lta) * 0.0075)
        p8_tot_ded = p8_pf + p8_esic + pt_p8 + lwf_p8
        p8_net = p8_gross - p8_tot_ded
        return p8_net

    # Write data with formulas
    for idx, card in enumerate(sorted(assoc_cards, key=lambda x: (x.get('state',''), x.get('city',''))), 1):
        r = idx + 1  # row number
        rs = str(r)

        state = card.get('state', '')
        city = card.get('city', '')
        site_code = card.get('site_codes', '')
        entity = card.get('entity', 'AMZL')
        pt_key = f"{state.upper()}|{city.upper()}"
        pt_data = alfa_pt_lwf.get(pt_key, {})

        # Determine if this site uses ROUND for PF (instead of ROUNDUP)
        use_round_pf = site_code.upper() in round_pf_sites
        pf_func = "ROUND" if use_round_pf else "ROUNDUP"

        # Determine if INFC + MH/GJ (or HMH4) for Holiday Pay formula override
        # Uses fixed 45 weekly hrs: PH10 = ((B+F+L)*12)/(52*45)*2*9, PH5 = *2*4
        is_holiday_override = (
            (entity.upper() == 'INFC' and state.upper() in ('MH', 'GJ', 'MAHARASHTRA', 'GUJARAT'))
            or site_code.upper() == 'HMH4'
        )

        # Static values (A-N)
        ws.cell(row=r, column=1, value=entity)
        ws.cell(row=r, column=2, value=state)
        ws.cell(row=r, column=3, value=city)
        ws.cell(row=r, column=4, value=site_code)
        ws.cell(row=r, column=5, value=card.get('daily_hours', 8))
        ws.cell(row=r, column=6, value=card.get('weekly_hours', 40))
        ws.cell(row=r, column=7, value="Y" if card.get('weekly_hours', 40) == 40 else "N")
        ws.cell(row=r, column=8, value=card.get('minimum_wage', 0))
        ws.cell(row=r, column=9, value=card.get('basic', 0))
        ws.cell(row=r, column=10, value=card.get('flexi', 0))
        ws.cell(row=r, column=11, value=card.get('lta', 0))
        ws.cell(row=r, column=12, value=card.get('hra', 0))
        ws.cell(row=r, column=13, value=f"=I{rs}+J{rs}+K{rs}+L{rs}")  # Gross
        ws.cell(row=r, column=14, value=15)  # Premium %

        # Premium columns (O-S) - formulas
        ws.cell(row=r, column=15, value=f"=ROUND(I{rs}*(1+N{rs}/100),0)")  # Basic_P
        ws.cell(row=r, column=16, value=f"=ROUND(J{rs}*(1+N{rs}/100),0)")  # Flexi_P
        ws.cell(row=r, column=17, value=f"=ROUND(K{rs}*(1+N{rs}/100),0)")  # LTA_P
        ws.cell(row=r, column=18, value=f"=ROUND(L{rs}*(1+N{rs}/100),0)")  # HRA_P
        ws.cell(row=r, column=19, value=f"=O{rs}+P{rs}+Q{rs}+R{rs}")  # Gross_P

        # P10 (T-AF) - formulas
        ws.cell(row=r, column=20, value=f"=ROUND(O{rs}/22,0)")  # P10_Basic
        ws.cell(row=r, column=21, value=f"=ROUND(P{rs}/22,0)")  # P10_Flexi
        ws.cell(row=r, column=22, value=f"=ROUND(Q{rs}/22,0)")  # P10_LTA
        ws.cell(row=r, column=23, value=f"=ROUND(R{rs}/22,0)")  # P10_HRA
        ws.cell(row=r, column=24, value=f"=T{rs}+U{rs}+V{rs}+W{rs}")  # P10_Gross
        # P10_OT - check for hardcoded override (smart bypass)
        p10_ot_hardcode = hardcoded_p10_ot.get((site_code.upper(), 'P10'))
        if p10_ot_hardcode:
            # Compute formula-based OT for smart bypass check
            basic_p = round(card.get('basic', 0) * 1.15)
            flexi_p = round(card.get('flexi', 0) * 1.15)
            lta_p = round(card.get('lta', 0) * 1.15)
            included_p = basic_p + flexi_p + lta_p
            wh = card.get('weekly_hours', 45)
            formula_ot = round(((included_p) * 12 / (52 * wh)) * 2) if wh == 40 else 0
            if formula_ot > p10_ot_hardcode:
                # Formula OT exceeds hardcode — bypass, use formula
                ws.cell(row=r, column=25, value=f'=ROUND(IF(G{rs}="Y",((O{rs}+P{rs}+Q{rs})*12/(52*F{rs}))*2,0),0)')
                hardcode_notifications.append(f"{site_code} P10_OT: Formula ({formula_ot}) > Hardcoded ({p10_ot_hardcode}) — using formula")
            else:
                ws.cell(row=r, column=25, value=p10_ot_hardcode)  # Hardcoded P10 OT
        else:
            ws.cell(row=r, column=25, value=f'=ROUND(IF(G{rs}="Y",((O{rs}+P{rs}+Q{rs})*12/(52*F{rs}))*2,0),0)')  # P10_OT formula

        # P10_Total Gross - always formula (Gross + OT)
        ws.cell(row=r, column=26, value=f"=X{rs}+Y{rs}")  # P10_Total
        ws.cell(row=r, column=27, value=f"={pf_func}(T{rs}*12%,0)")  # P10_PF (ROUND or ROUNDUP)
        ws.cell(row=r, column=28, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(T{rs}+U{rs}+V{rs})*0.75%),0)")  # P10_ESIC
        ws.cell(row=r, column=29, value=pt_data.get('pt_p10', 0))  # P10_PT
        ws.cell(row=r, column=30, value=pt_data.get('lwf_p10', 0))  # P10_LWF
        ws.cell(row=r, column=31, value=f"=AA{rs}+AB{rs}+AC{rs}+AD{rs}")  # P10_TotDed

        # P10_Net - always formula (Net = Total Gross - Deductions)
        ws.cell(row=r, column=32, value=f"=Z{rs}-AE{rs}")  # P10_Net

        # P5 (AG-AQ) - formulas
        ws.cell(row=r, column=33, value=f"=ROUND(O{rs}/22/E{rs}*4,0)")  # P5_Basic
        ws.cell(row=r, column=34, value=f"=ROUND(P{rs}/22/E{rs}*4,0)")  # P5_Flexi
        ws.cell(row=r, column=35, value=f"=ROUND(Q{rs}/22/E{rs}*4,0)")  # P5_LTA
        ws.cell(row=r, column=36, value=f"=ROUND(R{rs}/22/E{rs}*4,0)")  # P5_HRA
        ws.cell(row=r, column=37, value=f"=AG{rs}+AH{rs}+AI{rs}+AJ{rs}")  # P5_Gross
        ws.cell(row=r, column=38, value=f"={pf_func}(AG{rs}*12%,0)")  # P5_PF (ROUND or ROUNDUP)
        ws.cell(row=r, column=39, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(AG{rs}+AH{rs}+AI{rs})*0.75%),0)")  # P5_ESIC
        ws.cell(row=r, column=40, value=pt_data.get('pt_p5', 0))  # P5_PT
        ws.cell(row=r, column=41, value=pt_data.get('lwf_p5', 0))  # P5_LWF
        ws.cell(row=r, column=42, value=f"=AL{rs}+AM{rs}+AN{rs}+AO{rs}")  # P5_TotDed

        # P5_Net - always formula
        ws.cell(row=r, column=43, value=f"=AK{rs}-AP{rs}")  # P5_Net

        # P8 (AR-BB) - formulas
        ws.cell(row=r, column=44, value=f"=ROUND(O{rs}/22/E{rs}*8,0)")  # P8_Basic
        ws.cell(row=r, column=45, value=f"=ROUND(P{rs}/22/E{rs}*8,0)")  # P8_Flexi
        ws.cell(row=r, column=46, value=f"=ROUND(Q{rs}/22/E{rs}*8,0)")  # P8_LTA
        ws.cell(row=r, column=47, value=f"=ROUND(R{rs}/22/E{rs}*8,0)")  # P8_HRA
        ws.cell(row=r, column=48, value=f"=AR{rs}+AS{rs}+AT{rs}+AU{rs}")  # P8_Gross
        ws.cell(row=r, column=49, value=f"={pf_func}(AR{rs}*12%,0)")  # P8_PF (ROUND or ROUNDUP)
        ws.cell(row=r, column=50, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(AR{rs}+AS{rs}+AT{rs})*0.75%),0)")  # P8_ESIC
        ws.cell(row=r, column=51, value=pt_data.get('pt_p8', 0))  # P8_PT
        ws.cell(row=r, column=52, value=pt_data.get('lwf_p8', 0))  # P8_LWF
        ws.cell(row=r, column=53, value=f"=AW{rs}+AX{rs}+AY{rs}+AZ{rs}")  # P8_TotDed

        # P8_Net - always formula
        ws.cell(row=r, column=54, value=f"=AV{rs}-BA{rs}")  # P8_Net

        # Holiday (BC-BK) - formulas
        ws.cell(row=r, column=55, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(Z{rs}+X{rs})*0.75%),0)")  # PH10_ESIC
        ws.cell(row=r, column=56, value=f"=Z{rs}+BI{rs}-AA{rs}-AC{rs}-AD{rs}-BC{rs}")  # PH10_Net
        ws.cell(row=r, column=57, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(AK{rs}+AK{rs})*0.75%),0)")  # PH5_ESIC
        ws.cell(row=r, column=58, value=f"=AK{rs}+BJ{rs}-AL{rs}-AC{rs}-AD{rs}-BE{rs}")  # PH5_Net
        ws.cell(row=r, column=59, value=f"=ROUNDUP(IF((O{rs}+P{rs}+Q{rs})>21000,0,(AV{rs}+AV{rs})*0.75%),0)")  # PH8_ESIC
        ws.cell(row=r, column=60, value=f"=AV{rs}+BK{rs}-AW{rs}-AY{rs}-AZ{rs}-BG{rs}")  # PH8_Net

        # PH10, PH5, PH8 - INFC MH/GJ + HMH4 uses fixed 45 weekly hrs formula
        if is_holiday_override:
            ws.cell(row=r, column=61, value=f"=ROUND(((O{rs}+P{rs}+Q{rs})*12)/(52*45)*2*9,0)")  # PH10
            ws.cell(row=r, column=62, value=f"=ROUND(((O{rs}+P{rs}+Q{rs})*12)/(52*45)*2*4,0)")  # PH5
        else:
            ws.cell(row=r, column=61, value=f"=ROUND(((O{rs}+P{rs}+Q{rs})*12/(52*F{rs}))*E{rs},0)")  # PH10
            ws.cell(row=r, column=62, value=f"=ROUND(((O{rs}+P{rs}+Q{rs})*12/(52*F{rs}))*4,0)")  # PH5
        ws.cell(row=r, column=63, value=f"=BJ{rs}*2")  # PH8 = PH5 * 2

        # Apply colors and borders to this row
        for col in range(1, 64):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            if col in NET_COLS:
                cell.fill = FILL_NET
            elif col in GROSS_COLS:
                cell.fill = FILL_GROSS
            elif col in COMP_COLS:
                cell.fill = FILL_COMP

    # --- Remarks/Reference Sheet ---
    ws_remarks = wb.create_sheet("Remarks & Reference")
    remarks = [
        "ALFA RATE CARD — REMARKS & REFERENCE",
        "=" * 60,
        "",
        "1. HARD-CODED P10 OT SITES:",
        "   The following sites have P10_OT hard-coded (not formula-driven):",
        "   Net Pay flows naturally: Net = (P10_Gross + P10_OT) - Deductions",
        "   - UDL6 (UFF, Noida): P10_OT = 227",
        "   - LKOI, IXDD, KNUD, KNUO, AGRD, GKPL, MREE, VNSD, LKOA, LKOD (AMZL): P10_OT = 211",
        "   - LKO1 (INFC, Lucknow): P10_OT = 211",
        "   - LKOO (ATS, Lucknow): P10_OT = 211",
        "   - NCRJ, NCT3, NCT8, NZMF, NZMM (AMZL, Ghaziabad/Noida): P10_OT = 227",
        "",
        "2. ROUND PF SITES (instead of ROUNDUP):",
        "   - FHYE (UFF, Hyderabad): P10 Net = 1001, P5 Net = 386, P8 Net = 775",
        "   - SBLZ, UBL6, UBL9, UBL5, SBLY (UFF, Bangalore): P10 Net = 1291",
        "",
        "3. HOLIDAY PAY FORMULA OVERRIDE (INFC MH/GJ + HMH4):",
        "   Standard formula: PH10 = ((Basic_P+Flexi_P+LTA_P)*12/(52*WeeklyHrs))*DailyHrs",
        "   Override formula:  PH10 = ((Basic_P+Flexi_P+LTA_P)*12)/(52*45)*2*9",
        "                      PH5  = ((Basic_P+Flexi_P+LTA_P)*12)/(52*45)*2*4",
        "                      PH8  = PH5 * 2",
        "   Applies to: INFC sites in MH & GJ states + HMH4 (GSF HUB)",
        "",
        "4. SMART HARDCODE BYPASS:",
        "   If Gross increases and formula P10_OT > hardcoded P10_OT,",
        "   the hardcode is bypassed and formula is used instead.",
    ]

    # Add bypass notifications if any
    if hardcode_notifications:
        remarks.append("")
        remarks.append("   ⚠️ HARDCODE BYPASSED THIS EXPORT:")
        for notif in hardcode_notifications:
            remarks.append(f"   → {notif}")
    else:
        remarks.append("   No bypasses triggered this export.")

    remarks.extend([
        "",
        "5. PREMIUM: 15% applied to all components (Basic, Flexi, LTA, HRA)",
        "",
        "6. PF FORMULA: ROUNDUP(Basic*12%, 0) for most sites, ROUND for sites listed in #2",
        "",
        "7. ESIC: ROUNDUP applied; IF(Included > 21000, 0, (components)*0.75%)",
        "",
        "8. OT: Applicable only for Mandatory OT = Y (40hr sites)",
        "   Formula: ROUND(((Basic_P+Flexi_P+LTA_P)*12/(52*WeeklyHrs))*2, 0)",
        "",
        "=" * 60,
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        "Developed by: Ravi Kumar (Kmarnuz) | Sr. SME CTK MHLS",
    ])
    for i, line in enumerate(remarks, 1):
        ws_remarks.cell(row=i, column=1, value=line)
    ws_remarks.column_dimensions['A'].width = 80

    # Column widths
    for col in range(1, 64):
        ws.column_dimensions[get_column_letter(col)].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"ALFA_Rate_Card_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    save_audit_entry("DOWNLOAD", filename=fname, details=f"ALFA Rate Card with formulas ({len(assoc_cards)} sites)")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# --- Attendance Incentive Depository ---
AI_DEPOSITORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ai_depository.json')

def load_ai_depository():
    """Load AI depository from JSON file."""
    if os.path.exists(AI_DEPOSITORY_FILE):
        with open(AI_DEPOSITORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def apply_ai_depository_to_cards():
    """Apply AI + Region from depository to all matching cards (all year bands)."""
    dep = load_ai_depository()
    if not dep:
        return 0
    all_cards = db.list_wage_cards()
    updated = 0
    for card in all_cards:
        entity = card.get('entity', '')
        site = card.get('site_codes', '')
        short_bt = card.get('short_bt', '')
        key = f"{entity}|{site}|{short_bt}"
        if key in dep:
            info = dep[key]
            card['attendance_incentive'] = info['ai']
            card['region'] = info['region']
            db.put_wage_card(card, skip_save=True)
            updated += 1
    if updated > 0:
        db.save()
    return updated

@app.post("/api/ai-depository/upload")
async def upload_ai_depository(file: UploadFile = File(...), password: str = Form("")):
    """Upload new AI Depository Excel. Requires password. Auto-applies to all cards."""
    if password != get_upload_password():
        raise HTTPException(403, "Invalid password.")
    import openpyxl
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Parse: Entity, State, City, Site Code, MW Zone, Region, MW Category, Short BT, Tenure, AI
    dep = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 10:
            continue
        entity = str(row[0] or '').strip()
        site_code = str(row[3] or '').strip()
        region = str(row[5] or '').strip()
        short_bt = str(row[7] or '').strip()
        ai = row[9]
        if not entity or not site_code or not short_bt:
            continue
        try:
            ai = int(float(ai)) if ai else 0
        except (ValueError, TypeError):
            ai = 0
        key = f"{entity}|{site_code}|{short_bt}"
        dep[key] = {"ai": ai, "region": region}

    # Save depository
    with open(AI_DEPOSITORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(dep, f, indent=2)

    # Apply to all cards
    updated = apply_ai_depository_to_cards()
    save_audit_entry("UPLOAD", filename=file.filename, details=f"AI Depository: {len(dep)} entries, applied to {updated} cards")
    return {"status": "success", "entries": len(dep), "cards_updated": updated}

@app.get("/api/ai-depository/download")
def download_ai_depository():
    """Download current AI Depository as Excel."""
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill
    dep = load_ai_depository()
    if not dep:
        raise HTTPException(404, "No AI Depository found.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Depository"
    headers = ["Entity", "Site Code", "Short BT", "Attendance Incentive", "Region"]
    hdr_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin

    for idx, (key, val) in enumerate(sorted(dep.items()), 2):
        parts = key.split("|")
        ws.cell(row=idx, column=1, value=parts[0]).border = thin
        ws.cell(row=idx, column=2, value=parts[1]).border = thin
        ws.cell(row=idx, column=3, value=parts[2]).border = thin
        ws.cell(row=idx, column=4, value=val.get('ai', 0)).border = thin
        ws.cell(row=idx, column=5, value=val.get('region', '')).border = thin

    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=AI_Depository_Current.xlsx"})

@app.get("/api/ai-depository/status")
def ai_depository_status():
    """Get AI depository stats."""
    dep = load_ai_depository()
    return {"entries": len(dep), "file_exists": os.path.exists(AI_DEPOSITORY_FILE)}

@app.get("/", response_class=HTMLResponse)
def serve_frontend():
    return FRONTEND_HTML


if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*60)
    print("  WAGE CARD MANAGEMENT SYSTEM")
    print("  Open in browser: http://localhost:8000")
    print("="*60 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=8000, timeout_keep_alive=120)

