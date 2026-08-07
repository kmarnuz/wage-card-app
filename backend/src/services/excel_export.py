"""
Excel Export Service
Generates wage card Excel files matching the existing format.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions for the export
HEADER_COLUMNS = [
    ("Entity", "entity"),
    ("State", "state"),
    ("City", "city"),
    ("Site Code", "site_codes"),
    ("MW Zone", "mw_zone"),
    ("Region", "region"),
    ("MW Category", "mw_category"),
    ("Short BT", "short_bt"),
    ("Tenure", "tenure_years"),
    ("Weekly Hours", "weekly_hours"),
    ("Daily Hours", "daily_hours"),
    ("Minimum Wage", "minimum_wage"),
    ("MW Effective From", "mw_effective_date"),
    ("Basic", "basic"),
    ("Flexi Allowance", "flexi"),
    ("LTA", "lta"),
    ("HRA", "hra"),
    ("Conveyance Allowance", "conveyance"),
    ("Gross", "gross"),
    ("Per Hour OT Amount", "per_hour_ot_total"),
    ("PF (Employee)", "pf_employee"),
    ("ESIC (Employee)", "esic_employee"),
    ("Professional Tax", "pt_employee"),
    ("LWF (Employee)", "lwf_employee"),
    ("Gross Deductions", "gross_deductions"),
    ("Net Salary", "net_salary"),
    ("PF (Employer)", "pf_employer"),
    ("ESIC (Employer)", "esic_employer"),
    ("LWF (Employer)", "lwf_employer"),
    ("CTC", "ctc"),
    ("OT (Default)", "ot_default"),
    ("NSA", "nsa"),
    ("Attendance Incentive", "attendance_incentive"),
    ("Total Remuneration", "total_remuneration"),
    ("Excluded Wages", "excluded_wages"),
    ("50% Cap Amount", "cap_50_amount"),
    ("50% Cap (Met/Not Met)", "cap_50_met"),
    ("MW Compliant", "mw_compliant"),
    ("Hol Wage", "hol_wage"),
    ("Old Per Hr OT", "old_ot"),
    ("Old Hol Wage", "old_hol"),
    ("Balancing Pay OT", "bal_pay_ot"),
    ("Balancing Pay Hol Wage", "bal_pay_hol"),
]

# Styling
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
CURRENCY_FORMAT = '#,##0'
PERCENT_FORMAT = '0.00%'

# Highlight fills for special columns
FILL_NET = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green - Net Pay
FILL_GROSS = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # Light yellow - Gross
FILL_FLEXI = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue - Flexi/components

# Columns to highlight
HIGHLIGHT_MAP = {
    "net_salary": FILL_NET,
    "gross": FILL_GROSS,
    "ctc": FILL_NET,
    "flexi": FILL_FLEXI,
    "basic": FILL_FLEXI,
    "lta": FILL_FLEXI,
}
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def export_wage_cards_to_excel(cards: list[dict], group_by_tenure: bool = True) -> bytes:
    """
    Export wage cards to an Excel file (optimized for cloud free-tier).

    Args:
        cards: List of wage card dictionaries
        group_by_tenure: If True, creates separate sections for each tenure band

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    wb.calculation_on_load = True
    ws = wb.active
    ws.title = "Wage Cards"

    if group_by_tenure:
        # Sort cards by state, city, role, tenure
        cards = sorted(cards, key=lambda c: (
            c.get("state", ""),
            c.get("city", ""),
            c.get("short_bt", ""),
            c.get("tenure_years", 0),
        ))

    # Write headers (only headers get styling)
    for col_idx, (header_name, _) in enumerate(HEADER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Build column letter lookup
    cl = {}
    for idx, (_, fk) in enumerate(HEADER_COLUMNS, 1):
        cl[fk] = get_column_letter(idx)

    # Write data rows with Excel FORMULAS for calculated fields
    # NOTE: Styling removed from data rows for performance on free-tier hosting
    for row_idx, card in enumerate(cards, 2):
        r = str(row_idx)

        for col_idx, (_, field_key) in enumerate(HEADER_COLUMNS, 1):
            formula = None

            if field_key == "gross":
                formula = f"={cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r}+{cl['hra']}{r}+{cl['conveyance']}{r}"
            elif field_key == "per_hour_ot_total":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value=0)
                    continue
                formula = f"=ROUND((({cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r})*12/(52*{cl['weekly_hours']}{r}))*2,0)"
            elif field_key == "pf_employee":
                formula = f"=MIN(12%*{cl['basic']}{r},1800)"
            elif field_key == "esic_employee":
                incl = f"{cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r}"
                formula = f"=IF({incl}>21000,0,ROUNDUP(({incl})*0.75/100,0))"
            elif field_key == "gross_deductions":
                formula = f"={cl['pf_employee']}{r}+{cl['esic_employee']}{r}"
            elif field_key == "net_salary":
                formula = f"={cl['gross']}{r}-{cl['gross_deductions']}{r}"
            elif field_key == "pf_employer":
                formula = f"=MIN(13%*{cl['basic']}{r},1950)"
            elif field_key == "esic_employer":
                incl = f"{cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r}"
                formula = f"=IF({incl}>21000,0,ROUNDUP(({incl})*3.25/100,0))"
            elif field_key == "ctc":
                formula = f"={cl['gross']}{r}+{cl['pf_employer']}{r}+{cl['esic_employer']}{r}"
            elif field_key == "ot_default":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value=0)
                    continue
                formula = f"=IF({cl['weekly_hours']}{r}=40,{cl['per_hour_ot_total']}{r}*22,0)"
            elif field_key == "included_wages":
                formula = f"={cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r}"
            elif field_key == "total_remuneration":
                formula = f"={cl['ctc']}{r}+{cl['ot_default']}{r}+{cl['nsa']}{r}+{cl['attendance_incentive']}{r}"
            elif field_key == "included_pct":
                formula = f"=IF({cl['total_remuneration']}{r}>0,{cl['included_wages']}{r}/{cl['total_remuneration']}{r},0)"
            elif field_key == "excluded_wages":
                formula = f"={cl['ot_default']}{r}+{cl['nsa']}{r}+{cl['attendance_incentive']}{r}+{cl['hra']}{r}+{cl['conveyance']}{r}"
            elif field_key == "cap_50_amount":
                formula = f"=50%*{cl['total_remuneration']}{r}"
            elif field_key == "cap_50_met":
                formula = f'=IF({cl["excluded_wages"]}{r}<={cl["cap_50_amount"]}{r},"Met","Not Met")'
            elif field_key == "mw_compliant":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value="N/A")
                    continue
                incl = f"{cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r}"
                formula = f'=IF({incl}>={cl["minimum_wage"]}{r},"Yes","No")'
            elif field_key == "hol_wage":
                formula = f"=ROUND(({cl['basic']}{r}+{cl['flexi']}{r}+{cl['lta']}{r})*12/(52*{cl['weekly_hours']}{r})*1*{cl['daily_hours']}{r},0)"
            elif field_key == "old_ot":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    value = round(card.get("old_ot", 0) or 0)
                    ws.cell(row=row_idx, column=col_idx, value=value if value else "")
                continue
            elif field_key == "old_hol":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    value = round(card.get("old_hol", 0) or 0)
                    ws.cell(row=row_idx, column=col_idx, value=value if value else "")
                continue
            elif field_key == "bal_pay_ot":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=f"=MAX(0,{cl['old_ot']}{r}-{cl['per_hour_ot_total']}{r})")
                continue
            elif field_key == "bal_pay_hol":
                if card.get("is_pt") or "PT" in str(card.get("short_bt", "")):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=f"=MAX(0,{cl['old_hol']}{r}-{cl['hol_wage']}{r})")
                continue

            if formula:
                ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                ws.cell(row=row_idx, column=col_idx, value=card.get(field_key, ""))

    # Set fixed column widths (faster than auto-fit)
    for col_idx, (header_name, _) in enumerate(HEADER_COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header_name) + 2, 12)

    # Freeze panes: Row 1 (header) stays on top, Column A stays on left
    ws.freeze_panes = "B2"

    # Add a "Checks & Logic" sheet
    ws_checks = wb.create_sheet("Checks & Logic")
    checks = [
        "GB WAGE CARD MANAGEMENT SYSTEM — COMPLETE LOGIC REFERENCE",
        "================================================================",
        "Developed by: Ravi Kumar (Kmarnuz) | Sr. SME CTK MHLS",
        "",
        "GROSS SPLIT LOGIC (All States except MH/WB):",
        "  1. Basic = MIN(Gross, 15,000)",
        "  2. Flexi = MIN(remaining, 7,500)",
        "  3. LTA = ONLY if MW > 22,500 then LTA = MW - 22,500",
        "  4. HRA = MIN(remaining, 7,500)",
        "  5. Conveyance = residual",
        "",
        "GROSS SPLIT LOGIC (MH and WB States):",
        "  1. Basic = MIN(Gross, 15,000)",
        "  2. HRA = 5% of MW (mandatory minimum, cap 7,500)",
        "  3. Flexi = MIN(remaining, 7,500)",
        "  4. LTA = ONLY if MW > 22,500 then LTA = MW - 22,500",
        "  5. Conveyance = residual",
        "",
        "WAGE CLASSIFICATION:",
        "  Included Wages = Basic + Flexi + LTA (Statutory Base)",
        "  Excluded Wages = HRA + Conveyance + OT + NSA + Incentive",
        "",
        "STATUTORY FORMULAS:",
        "  PF (Employee) = MIN(12% x Basic, 1,800)",
        "  PF (Employer) = MIN(13% x Basic, 1,950)",
        "  ESIC (Employee) = IF(Included > 21,000, 0, ROUNDUP(Included x 0.75%))",
        "  ESIC (Employer) = IF(Included > 21,000, 0, ROUNDUP(Included x 3.25%))",
        "  Per Hour OT = ROUND((Basic+Flexi+LTA) x 12 / (52 x Weekly Hrs) x 2)",
        "  OT Default = IF(Weekly Hrs = 40, Per Hour OT x 22, 0)",
        "  Hol Wage = ROUND((Basic+Flexi+LTA) x 12 / (52 x Weekly Hrs) x 1 x Daily Hrs)",
        "",
        "SALARY CALCULATIONS:",
        "  Gross = Basic + Flexi + LTA + HRA + Conveyance",
        "  Gross Deductions = PF (Employee) + ESIC (Employee)",
        "  Net Salary = Gross - Gross Deductions",
        "  CTC = Gross + PF (Employer) + ESIC (Employer)",
        "  Total Remuneration = CTC + Default OT + NSA + Attendance Incentive",
        "",
        "COMPLIANCE CHECKS:",
        "  MW Compliant = Basic + Flexi + LTA >= Minimum Wage",
        "  50% Cap = Excluded Wages <= 50% of Total Remuneration",
        "",
        "MW ZONE CLASSIFICATION:",
        "  Same city can have different MW for different zones",
        "  MW lookup: State + City + Short BT + Zone + MW Category",
        "",
        "MW CATEGORY:",
        "  Associate and Associate PT = Semi Skilled",
        "  All other roles = Skilled",
        "",
        "COMPONENT CAPS:",
        "  Basic: 15,000 max | Flexi: 7,500 max | HRA: 7,500 max",
        "  LTA: Only when MW > 22,500 | PF EE: 1,800 | PF ER: 1,950",
        "  ESIC Ceiling: 21,000",
        "",
        "BALANCING PAY:",
        "  Old Per Hr OT and Old Hol Wage from template (reference)",
        "  Bal Pay OT = MAX(0, Old OT - New Per Hr OT)",
        "  Bal Pay Hol = MAX(0, Old Hol - New Hol Wage)",
        "  PT cards: Blank (not applicable)",
        "",
        "PT and LWF: Shown as As applicable (not deducted from Net)",
        "",
        "================================================================",
        "ASSOCIATE PT WAGE CARD LOGIC:",
        "================================================================",
        "Auto-derived from Associate cards (0-4 Year). OT=0. MW=N/A.",
        "",
        "ENTITY RULES:",
        "  INFC BLR7/DED3:  PT Gross = 50% of Associate Gross",
        "  INFC BOM5/PNQ3:  PT Gross = 65% of Associate Gross",
        "  AMZL (All):      PT Net = ROUNDUP(65% x (Assoc Net + Default OT))",
        "  UFF (All):       PT Net = ROUNDUP(65% x (Assoc Net + Default OT))",
        "  ATS (MH state):  PT Gross = 65% of Associate Gross",
        "  ATS (40hr):      PT Gross = 63% of Associate Gross",
        "  ATS (others):    PT Gross = 50% of Associate Gross",
        "  AMXL (40hr):     PT Gross = 65% of Associate Gross",
        "  AMXL (others):   PT Gross = 50% of Associate Gross",
        "",
        "PT SPLIT (MH/WB):",
        "  Basic = PT Gross / 1.05 (cap 15,000)",
        "  HRA = PT Gross - Basic (remainder absorbs rounding)",
        "  Flexi = 0 | Conveyance = 0",
        "  So: Basic + HRA = PT Gross exactly",
        "",
        "PT SPLIT (All other states):",
        "  Standard split: Basic -> Flexi -> HRA -> Conveyance",
        "  MW not enforced (MW=0 for split purposes)",
        "",
        "PT SPECIFIC RULES:",
        "  Per Hour OT = 0 | OT Default = 0 | MW Compliance = N/A",
        "  Bal Pay OT/Hol = Blank",
        "",
        "================================================================",
        "MW REVISION LOGIC:",
        "================================================================",
        "  If New MW <= 0Yr Gross: Keep Gross, re-split only",
        "  If New MW > 0Yr Gross:",
        "    Gap = New MW - 0Yr Gross",
        "    All years increased by Gap (0Yr=MW, 1-4Yr=old+Gap)",
        "    Then re-split and recalculate all",
        "",
        "================================================================",
        "TEMPLATE FORMAT (Single Tab):",
        "================================================================",
        "  Mile | Site Code | State | City | Weekly Hrs | Daily Hrs |",
        "  Node | Level | Short BT | MW Category | MW Zone |",
        "  Minimum Wage | MW Effective Date |",
        "  0 Year | 1 Year | 2 Year | 3 Year | 4 Year |",
        "  Old OT 0-4Yr | Old Hol 0-4Yr",
        "",
        "================================================================",
        "PARITY GROUPS:",
        "================================================================",
        "Sites in a parity group maintain IDENTICAL Gross and component split",
        "regardless of MW zone differences. Highest Gross in group is applied to all.",
        "",
        "  1. UFF Bangalore: SBLY, SBLZ, UBL5, UBL6, UBL9",
        "     (SBLZ Zone 2 follows Zone 1 rates)",
        "  2. UFF Delhi-NCR: PDL1, PDL2, PDL5, UDL4",
        "     (HY sites match DL sites)",
        "  3. AMZL Delhi-NCR-Faridabad: DELH, DLIH, FADA, NCT2, NZMN + DL sites",
        "     (HY MW 18,501 matches DL MW 22,411)",
        "",
        "================================================================",
        "IXCE ASSOCIATE PT HARDCODE:",
        "================================================================",
        "  IXCE (AMZL, Panchkula): Fixed PT Gross values per tenure year",
        "  0Yr=11,071 | 1Yr=11,499 | 2Yr=11,922 | 3Yr=12,133 | 4Yr=12,346",
        "  Smart Bypass: If formula > hardcode after MW revision, formula used.",
        "",
        "================================================================",
        "ROUNDING RULES:",
        "================================================================",
        "  All monetary values: Rounded to whole numbers (no decimals)",
        "  PF: ROUNDUP | ESIC: ROUNDUP | OT/Hol Wage: ROUND",
        "",
        "================================================================",
    ]
    for i, check in enumerate(checks, 1):
        ws_checks.cell(row=i, column=1, value=check)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
